# -*- coding: utf-8 -*-
"""
계좌현황 현재가(I/J) · 환율 · 평가액 수식 갱신
- Excel에서는 GOOGLEFINANCE가 동작하지 않음 → yfinance/네이버로 가격을 채워 넣음
- K/L/M/N은 Excel에서 계산 가능한 수식으로 유지 (환율은 J5 숫자 참조)
"""
import argparse
import re
import sys
import time
import warnings
from pathlib import Path

import openpyxl
import pandas as pd
import requests
import yfinance as yf

from app_paths import app_dir, resolve_workbook, safe_reconfigure_stdio

warnings.filterwarnings("ignore")
safe_reconfigure_stdio()

DEFAULT_FILE = app_dir() / "investment_0719.xlsx"
SHEET = "계좌현황"
HOLD_START = 9
HOLD_END_MAX = 120


def is_cash(tic, name):
    tic = str(tic or "")
    name = str(name or "")
    if tic == "현금":
        return True
    return any(k in name for k in ("예수금", "CMA", "MMF", "예금", "현금성", "새마을"))


def to_yf(tic, country):
    t = str(tic).strip()
    if country == "미국":
        return t
    if re.fullmatch(r"\d+(\.0+)?", t):
        return f"{str(int(float(t))).zfill(6)}.KS"
    if re.fullmatch(r"[0-9A-Za-z]{6}", t):
        return f"{t}.KS"
    return f"{t}.KS"


def naver_price(code: str):
    """한국 종목 현재가 (네이버)"""
    code = str(code).strip()
    if re.fullmatch(r"\d+(\.0+)?", code):
        code = str(int(float(code))).zfill(6)
    url = f"https://finance.naver.com/item/sise.naver?code={code}"
    try:
        r = requests.get(url, timeout=10, headers={"User-Agent": "Mozilla/5.0"})
        r.raise_for_status()
        # <strong id="_nowVal">12,345</strong> 또는 유사 패턴
        m = re.search(r'id="_nowVal"[^>]*>([\d,]+)<', r.text)
        if not m:
            m = re.search(r"now_value[^>]*>([\d,]+)<", r.text)
        if m:
            return float(m.group(1).replace(",", ""))
    except Exception:
        pass
    return None


def yf_price(symbol: str):
    try:
        t = yf.Ticker(symbol)
        info = t.fast_info
        px = getattr(info, "last_price", None) or getattr(info, "lastPrice", None)
        if px is None:
            hist = t.history(period="5d")
            if len(hist):
                px = float(hist["Close"].iloc[-1])
        return float(px) if px is not None else None
    except Exception:
        return None


def get_fx():
    px = yf_price("USDKRW=X")
    if px:
        return round(px, 2)
    # fallback
    try:
        r = requests.get(
            "https://query1.finance.yahoo.com/v8/finance/chart/USDKRW=X",
            timeout=10,
            headers={"User-Agent": "Mozilla/5.0"},
        )
        return round(r.json()["chart"]["result"][0]["meta"]["regularMarketPrice"], 2)
    except Exception:
        return 1488.0


def looks_korean_ticker(tic):
    t = str(tic or "").strip()
    if re.fullmatch(r"\d+(\.0+)?", t):
        return True
    # 한국 ETF 코드: 0008S0, 0072R0 등 (숫자+영문 혼합 6자리)
    if re.fullmatch(r"[0-9][0-9A-Za-z]{5}", t):
        return True
    return False


def detect_country(ws, r):
    tic = ws.cell(r, 4).value
    # 종목코드가 한국형이면 무조건 한국 (C열이 잘못되어 있어도)
    if looks_korean_ticker(tic):
        return "한국"
    c = ws.cell(r, 3).value
    if c in ("한국", "미국"):
        return c
    if isinstance(c, str) and c.startswith("="):
        # 수식이면 평단가 열로 추정
        pass
    elif isinstance(c, str) and "미국" in c:
        return "미국"
    h = ws.cell(r, 8).value  # 평단가 달러
    g = ws.cell(r, 7).value  # 평단가 원
    if isinstance(h, (int, float)) and h > 0 and not (isinstance(g, (int, float)) and g > 0):
        return "미국"
    return "한국"


def last_hold_row(ws):
    last = HOLD_START
    for r in range(HOLD_START, HOLD_END_MAX + 1):
        tic = ws.cell(r, 4).value
        name = ws.cell(r, 5).value
        if tic is None and name is None:
            continue
        if str(tic or "").strip() == "" and str(name or "").strip() == "":
            break
        last = r
    return last


def update_workbook(file_path: Path):
    fx = get_fx()
    print(f"\n=== {file_path.name} ===")
    print(f"환율 USD/KRW = {fx}")

    wb = openpyxl.load_workbook(file_path)
    ws = wb[SHEET]
    hold_end = last_hold_row(ws)
    print(f"종목 행: {HOLD_START}–{hold_end}")

    # J5: 숫자 환율 (Excel 계산용). 표시는 I5에 문구
    ws.cell(5, 9).value = f"$1 = ￦{fx:,.0f}"
    ws.cell(5, 10).value = fx
    ws.cell(5, 10).number_format = "0.00"

    ok, fail = 0, []
    for r in range(HOLD_START, hold_end + 1):
        tic = ws.cell(r, 4).value
        name = ws.cell(r, 5).value
        if tic is None and name is None:
            continue

        if is_cash(tic, name):
            # 현금: I=G, J 비움
            ws.cell(r, 9).value = f"=G{r}"
            ws.cell(r, 10).value = None
        else:
            country = detect_country(ws, r)
            # 국가 열 보정 (수식이거나 코드와 불일치하면 값으로 고정)
            cval = ws.cell(r, 3).value
            if (
                (isinstance(cval, str) and cval.startswith("="))
                or cval not in ("한국", "미국")
                or (looks_korean_ticker(tic) and cval != "한국")
            ):
                ws.cell(r, 3).value = country

            px = None
            if country == "한국":
                px = naver_price(str(tic))
                if px is None:
                    px = yf_price(to_yf(tic, country))
                if px is not None:
                    ws.cell(r, 9).value = px  # I 원화
                    ws.cell(r, 9).number_format = "#,##0.##"
                    ws.cell(r, 10).value = None
                    ok += 1
                    print(f"  {r:3d} [KR] {tic} {str(name)[:24]:24} → {px:,.2f}")
                else:
                    fail.append((r, tic, name, "KR"))
            else:
                px = yf_price(to_yf(tic, country))
                if px is not None:
                    ws.cell(r, 9).value = None
                    ws.cell(r, 10).value = px  # J 달러
                    ws.cell(r, 10).number_format = "0.00"
                    ok += 1
                    print(f"  {r:3d} [US] {tic} {str(name)[:24]:24} → ${px:,.2f}")
                else:
                    fail.append((r, tic, name, "US"))
            time.sleep(0.05)

        # K/L/M/N: Excel에서 계산되는 수식 (환율 $J$5)
        ws.cell(r, 11).value = f"=IFERROR(I{r}*F{r}+J{r}*$J$5*F{r},)"
        ws.cell(r, 12).value = f"=IFERROR(K{r}/$K$8,)"
        ws.cell(r, 13).value = f"=IFERROR(K{r}-G{r}*F{r}-H{r}*$J$5*F{r},)"
        ws.cell(r, 14).value = f"=IFERROR(M{r}/(K{r}-M{r}),)"

    ws.cell(8, 11).value = f"=SUM(K{HOLD_START}:K{hold_end})"
    ws.cell(8, 13).value = f"=SUM(M{HOLD_START}:M{hold_end})"
    ws.cell(8, 14).value = "=IFERROR(M8/(K8-M8),)"

    wb.save(file_path)
    print(f"완료: 성공 {ok}건 / 실패 {len(fail)}건")
    for r, tic, name, c in fail:
        print(f"  FAIL {r} [{c}] {tic} {name}")
    print(f"저장: {file_path}")


def main():
    ap = argparse.ArgumentParser(description="계좌현황 현재가(I/J) 갱신")
    ap.add_argument(
        "files",
        nargs="*",
        help="xlsx 경로 (미지정 시 investment_0719.xlsx)",
    )
    args = ap.parse_args()
    base = app_dir()
    if args.files:
        paths = []
        for f in args.files:
            paths.append(resolve_workbook(Path(f)))
    else:
        paths = [DEFAULT_FILE]

    for p in paths:
        if not p.exists():
            print(f"파일 없음: {p}")
            continue
        update_workbook(p)
    print("\n엑셀을 다시 열면 I/J 현재가와 K 평가액이 반영됩니다.")
    print("예: python update_prices.py investment_0719_W.xlsx")


if __name__ == "__main__":
    main()
