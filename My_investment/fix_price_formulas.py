# -*- coding: utf-8 -*-
"""계좌현황 I~N열: __xludf 제거 → 실제 GOOGLEFINANCE/IMPORTXML 수식 복구."""
import sys
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

FILE = Path(r"c:\Users\seaho\My project\My_investment\investment_0719.xlsx")
SHEET = "계좌현황"
HOLD_START = 9
HOLD_END = 82


def is_cash_row(ws, r):
    tic = str(ws.cell(r, 4).value or "")
    name = str(ws.cell(r, 5).value or "")
    if tic == "현금":
        return True
    keys = ("예수금", "CMA", "MMF", "예금", "현금성", "새마을")
    return any(k in name for k in keys)


def formulas_for_row(r, cash: bool):
    """Return dict col -> formula string for I,J,K,L,M,N."""
    if cash:
        i = f"=G{r}"
        j = None
        m = f"=IFERROR(K{r}-G{r}*F{r}-H{r}*$J$5*F{r},)"
    else:
        i = (
            f'=IFERROR(IFERROR(IF($C{r}="한국",GOOGLEFINANCE($D{r}),""),'
            f"IMPORTXML('IMPORTXML함수'!$C$24&D{r},'IMPORTXML함수'!$C$25)),)"
        )
        j = f'=IFERROR(IF($C{r}="미국",GOOGLEFINANCE($D{r}),""),)'
        m = f"=IFERROR(K{r}-G{r}*F{r}-H{r}*$J$5*F{r},)"

    k = f'=IFERROR(I{r}*F{r}+J{r}*GOOGLEFINANCE("usdkrw")*F{r},)'
    l = f"=K{r}/$K$8"
    n = f"=IFERROR(M{r}/(K{r}-M{r}),)"
    return {9: i, 10: j, 11: k, 12: l, 13: m, 14: n}


def main():
    wb = openpyxl.load_workbook(FILE)
    ws = wb[SHEET]

    # FX display in J5
    ws.cell(5, 10).value = '="$1 = ￦"&ROUND(GOOGLEFINANCE("usdkrw"),0)'

    fixed = 0
    for r in range(HOLD_START, HOLD_END + 1):
        tic = ws.cell(r, 4).value
        name = ws.cell(r, 5).value
        if tic is None and name is None:
            continue

        cash = is_cash_row(ws, r)
        fml = formulas_for_row(r, cash)
        for col, val in fml.items():
            ws.cell(r, col).value = val
        fixed += 1

        # 국가 수식이 다른 행을 가리키면 현재 행으로 교정
        cval = ws.cell(r, 3).value
        if (
            not cash
            and isinstance(cval, str)
            and cval.startswith("=")
            and f"H{r}" not in cval
        ):
            ws.cell(r, 3).value = f'=IF(H{r}>0,"미국","한국")'
        elif not cash and cval is None:
            ws.cell(r, 3).value = f'=IF(H{r}>0,"미국","한국")'

    # K8 / M8 totals already OK; N8 depends on them
    ws.cell(8, 11).value = f"=SUM(K{HOLD_START}:K{HOLD_END})"
    ws.cell(8, 13).value = f"=SUM(M{HOLD_START}:M{HOLD_END})"
    ws.cell(8, 14).value = "=IFERROR(M8/(K8-M8),)"

    wb.save(FILE)
    print(f"fixed {fixed} rows in {SHEET} columns I-N")
    print("sample row 9:")
    for c, letter in [(9, "I"), (10, "J"), (11, "K"), (12, "L"), (13, "M"), (14, "N")]:
        print(f"  {letter}: {ws.cell(9, c).value}")
    print("sample cash row 77:")
    for c, letter in [(9, "I"), (10, "J"), (11, "K"), (13, "M")]:
        print(f"  {letter}: {ws.cell(77, c).value}")
    print("sample US row 41:")
    for c, letter in [(3, "C"), (9, "I"), (10, "J"), (11, "K")]:
        print(f"  {letter}: {ws.cell(41, c).value}")


if __name__ == "__main__":
    main()
