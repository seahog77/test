# -*- coding: utf-8 -*-
"""계좌현황 기반 월별 배당(실적+예측) → 마지막 탭 '월별배당' 생성."""
import calendar
import re
import sys
import warnings
from collections import defaultdict
from datetime import date
from pathlib import Path

import openpyxl
import pandas as pd
import yfinance as yf
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app_paths import app_dir, resolve_workbook, safe_reconfigure_stdio

warnings.filterwarnings("ignore")
safe_reconfigure_stdio()

DEFAULT_FILE = app_dir() / "investment_0719.xlsx"
SHEET_HOLD = "계좌현황"
SHEET_OUT = "월별배당"
YEAR = 2026
MIN_AMOUNT = 1000
HOLD_START = 9


def cached_number(v):
    if isinstance(v, (int, float)):
        return float(v)
    if not isinstance(v, str):
        return None
    m = re.search(r",\s*([-+]?\d+(?:\.\d+)?(?:[eE][-+]?\d+)?)\s*\)\s*$", v)
    return float(m.group(1)) if m else None


def cell_num(ws_d, ws_f, r, c):
    v = ws_d.cell(r, c).value
    if isinstance(v, (int, float)):
        return float(v)
    return cached_number(ws_f.cell(r, c).value) or 0.0


def get_fx(ws_d, ws_f=None):
    # 계좌현황 예: "$1 = ￦1488" (값 또는 수식 캐시)
    sheets = [ws_d] + ([ws_f] if ws_f is not None else [])
    for ws in sheets:
        for r in range(1, 10):
            for c in range(1, 15):
                v = ws.cell(r, c).value
                if not isinstance(v, str):
                    continue
                m = re.search(r"[￦]\s*([\d,]+)", v)
                if m:
                    return float(m.group(1).replace(",", ""))
    return 1488.0


def clean_name(name):
    n = str(name)
    for suffix in [
        "(DC)",
        "(ISA)",
        "(개인연금)",
        "(삼성개인연금)",
        "(미래에셋IRP)",
        ", DC",
    ]:
        n = n.replace(suffix, "")
    return n.strip()


# 상장 직후 등 이력이 짧아도 월배당으로 볼 종목 (한국 티커)
FORCE_MONTHLY_TICKERS = {
    "0219E0",  # KODEX 200커버드콜액티브
}


def looks_monthly_payer(ticker, name: str) -> bool:
    """월배당·커버드콜(위클리/데일리/타겟) 추정."""
    t = str(ticker or "").strip().upper()
    if t in FORCE_MONTHLY_TICKERS:
        return True
    n = str(name or "")
    if "월배당" in n:
        return True
    if "커버드콜" in n or "커버드" in n:
        return True
    if any(k in n for k in ("위클리", "데일리", "타켓커버드", "타겟커버드")):
        return True
    return False


def to_yf_ticker(ticker, country):
    t = str(ticker).strip()
    if country == "미국":
        return t
    # 한국 숫자/알파벳 혼합 코드 (0008S0, 0219E0 등)
    if re.fullmatch(r"\d{6}", t):
        return f"{t}.KS"
    if re.fullmatch(r"[0-9A-Za-z]{6}", t):
        return f"{t}.KS"
    if t.replace(".", "").isdigit():
        return f"{str(int(float(t))).zfill(6)}.KS"
    return f"{t}.KS"


def _strip_tz(index):
    """tz-aware DatetimeIndex → naive (버전별 API 호환)."""
    if getattr(index, "tz", None) is None:
        return index
    try:
        return index.tz_localize(None)
    except TypeError:
        return index.tz_convert("UTC").tz_localize(None)


def typical_payday(year, month, sample_dates):
    same_month = [d.day for d in sample_dates if d.month == month]
    if same_month:
        day = int(round(sum(same_month) / len(same_month)))
    else:
        day = 28 if month == 2 else 30 if month in (4, 6, 9, 11) else 31
    day = min(day, calendar.monthrange(year, month)[1])
    return date(year, month, day)


def looks_korean_ticker(tic):
    t = str(tic or "").strip()
    if re.fullmatch(r"\d+(\.0+)?", t):
        return True
    if re.fullmatch(r"[0-9][0-9A-Za-z]{5}", t):
        return True
    return False


def normalize_country(country, ticker, ws_f, r):
    if looks_korean_ticker(ticker):
        return "한국"
    if country in ("한국", "미국"):
        return country
    if isinstance(country, str) and "미국" in country:
        return "미국"
    if isinstance(country, str) and country.startswith("="):
        h = ws_f.cell(r, 8).value
        g = ws_f.cell(r, 7).value
        if isinstance(h, (int, float)) and h > 0 and not (
            isinstance(g, (int, float)) and g > 0
        ):
            return "미국"
    return "한국"


def load_holdings(path):
    wb_d = openpyxl.load_workbook(path, data_only=True)
    wb_f = openpyxl.load_workbook(path, data_only=False)
    ws_d = wb_d[SHEET_HOLD]
    ws_f = wb_f[SHEET_HOLD]
    fx = get_fx(ws_d, ws_f)

    rows = []
    for r in range(HOLD_START, 300):
        tic = ws_d.cell(r, 4).value
        name = ws_d.cell(r, 5).value
        if tic is None and name is None:
            if rows and r > rows[-1]["row"] + 2:
                break
            continue
        if name and "포트폴리오" in str(name):
            break
        country = ws_d.cell(r, 3).value or ws_f.cell(r, 3).value
        country = normalize_country(country, tic, ws_f, r)
        qty = cell_num(ws_d, ws_f, r, 6)
        val = cell_num(ws_d, ws_f, r, 11)
        if val <= 0:
            i_px = cell_num(ws_d, ws_f, r, 9)
            j_px = cell_num(ws_d, ws_f, r, 10)
            if qty > 0 and (i_px > 0 or j_px > 0):
                val = i_px * qty + j_px * fx * qty
        cat = ws_d.cell(r, 15).value or ""
        acct = ws_d.cell(r, 1).value or ""
        if val <= 0 and qty <= 0:
            continue
        rows.append(
            {
                "row": r,
                "acct": acct,
                "country": country,
                "ticker": tic,
                "name": name,
                "qty": qty,
                "val": val,
                "cat": cat,
            }
        )
    return rows, fx


def is_cash(row):
    tic = str(row["ticker"])
    name = str(row["name"])
    cat = str(row["cat"])
    if tic in ("현금", "예금") or cat in ("현금", "예금", "예수금"):
        return True
    if any(k in name for k in ("CMA", "MMF", "예수금", "예금", "현금성", "새마을")):
        return True
    return False


def build_records(holdings, fx, year=YEAR):
    # aggregate by ticker
    agg = defaultdict(lambda: {"qty": 0.0, "val": 0.0, "name": "", "country": "한국", "cat": "", "cash": False})
    for h in holdings:
        cash = is_cash(h)
        key = f"CASH::{h['name']}" if cash else f"TIC::{h['ticker']}"
        agg[key]["qty"] += h["qty"] or 0
        agg[key]["val"] += h["val"] or 0
        if h["val"] >= agg[key]["val"] - h["val"]:
            agg[key]["name"] = clean_name(h["name"])
            agg[key]["country"] = h["country"]
            agg[key]["cat"] = h["cat"]
            agg[key]["ticker"] = h["ticker"]
            agg[key]["cash"] = cash

    records = []
    detail = []

    for key, row in agg.items():
        name = row["name"]
        ticker = row.get("ticker", "")
        qty = row["qty"]
        val = row["val"]
        country = row["country"]
        cat = row["cat"]

        if row["cash"]:
            rate = 0.035 if any(k in name for k in ("CMA", "RP", "MMF")) else 0.03
            monthly_int = round(val * rate / 12)
            if monthly_int >= MIN_AMOUNT:
                for m in range(1, 13):
                    day = calendar.monthrange(year, m)[1]
                    records.append(
                        {
                            "date": date(year, m, day),
                            "ticker": ticker or "현금",
                            "name": name,
                            "krw": monthly_int,
                            "usd": 0.0,
                            "predict": True,
                            "month": m,
                            "kind": "이자",
                        }
                    )
                detail.append(
                    {
                        "ticker": ticker or "현금",
                        "name": name,
                        "freq": "이자(월)",
                        "annual": monthly_int * 12,
                        "yield": rate * 100,
                        "val": val,
                    }
                )
            continue

        yf_t = to_yf_ticker(ticker, country)
        try:
            divs = yf.Ticker(yf_t).dividends
            if len(divs) and getattr(divs.index, "tz", None) is not None:
                divs = divs.copy()
                divs.index = _strip_tz(divs.index)
        except Exception as e:
            print(f"  skip {yf_t}: {e}")
            divs = pd.Series(dtype=float)

        if len(divs) == 0:
            detail.append(
                {
                    "ticker": ticker,
                    "name": name,
                    "freq": "무배당/미확인",
                    "annual": 0,
                    "yield": 0,
                    "val": val,
                }
            )
            continue

        hist = divs[divs.index >= pd.Timestamp(f"{year - 2}-01-01")]
        hist_df = hist.reset_index()
        hist_df.columns = ["date", "amt"]
        hist_df["month"] = hist_df["date"].dt.month
        month_pattern = hist_df.groupby("month")["amt"].mean()
        sample_dates = list(hist_df["date"].dt.date)

        last6 = divs[divs.index >= pd.Timestamp(f"{year}-01-01") - pd.DateOffset(months=6)]
        force_monthly = looks_monthly_payer(ticker, name)
        if len(last6) >= 2:
            l6 = last6.reset_index()
            l6.columns = ["date", "amt"]
            l6["mk"] = l6["date"].dt.to_period("M")
            recent_ps = l6.groupby("mk")["amt"].sum().mean()
            is_monthly = force_monthly or len(l6.groupby("mk")) >= 8
            freq = "월배당" if is_monthly else "분기+"
        else:
            # 신규 상장 월배당(예: 0219E0) — 이력이 1건이어도 최근 주당배당으로 월 예측
            recent_ps = float(last6.iloc[-1]) if len(last6) else float(divs.iloc[-1])
            is_monthly = force_monthly
            freq = "월배당" if is_monthly else "연/반기"

        annual_est = 0.0
        first_hist_month = int(min(month_pattern.index)) if len(month_pattern) else None
        short_hist = force_monthly and len(month_pattern) < 3
        for m in range(1, 13):
            month_start = pd.Timestamp(f"{year}-{m:02d}-01")
            month_end = month_start + pd.DateOffset(months=1)
            actual = divs[(divs.index >= month_start) & (divs.index < month_end)]
            if len(actual) > 0:
                per_share = float(actual.sum())
                pay_date = actual.index[-1].date()
                is_predict = False
            elif m in month_pattern.index:
                per_share = float(month_pattern[m])
                pay_date = typical_payday(year, m, sample_dates)
                is_predict = True
            elif is_monthly and recent_ps > 0:
                # 신규 월배당(이력 짧음): 첫 배당월 이전은 예측하지 않음
                if short_hist and first_hist_month is not None and m < first_hist_month:
                    continue
                per_share = float(recent_ps)
                pay_date = typical_payday(year, m, sample_dates)
                is_predict = True
            else:
                continue

            if country == "미국":
                usd = round(per_share * qty, 4)
                krw = 0.0
                total_krw = usd * fx
            else:
                krw = round(per_share * qty)
                usd = 0.0
                total_krw = krw

            if total_krw < MIN_AMOUNT:
                continue

            annual_est += total_krw
            records.append(
                {
                    "date": pay_date,
                    "ticker": ticker,
                    "name": name,
                    "krw": krw,
                    "usd": usd,
                    "predict": is_predict,
                    "month": m,
                    "kind": "예측" if is_predict else "실적",
                }
            )

        detail.append(
            {
                "ticker": ticker,
                "name": name,
                "freq": freq,
                "annual": int(annual_est),
                "yield": round(annual_est / val * 100, 2) if val > 0 else 0,
                "val": val,
            }
        )

    # merge same ticker+month
    merged = {}
    for r in records:
        key = (str(r["ticker"]), r["month"], r["name"])
        if key not in merged:
            merged[key] = r.copy()
        else:
            merged[key]["krw"] += r["krw"]
            merged[key]["usd"] += r["usd"]
            # 실적 우선
            if not r["predict"]:
                merged[key]["predict"] = False
                merged[key]["kind"] = "실적"
                merged[key]["date"] = r["date"]

    final_records = sorted(merged.values(), key=lambda x: (x["date"], str(x["ticker"])))
    return final_records, detail, fx


def write_sheet(path, records, detail, fx, year=YEAR):
    wb = openpyxl.load_workbook(path)
    if SHEET_OUT in wb.sheetnames:
        del wb[SHEET_OUT]
    ws = wb.create_sheet(SHEET_OUT)  # last tab

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    predict_fill = PatternFill("solid", fgColor="FFF2CC")
    actual_fill = PatternFill("solid", fgColor="E2EFDA")
    thin = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )
    title_font = Font(bold=True, size=14, color="1F4E79")

    monthly = defaultdict(lambda: {"actual": 0.0, "predict": 0.0})
    for r in records:
        tot = r["krw"] + r["usd"] * fx
        if r["predict"]:
            monthly[r["month"]]["predict"] += tot
        else:
            monthly[r["month"]]["actual"] += tot

    annual = sum(m["actual"] + m["predict"] for m in monthly.values())

    ws["A1"] = f"{year}년 월별 배당·이자 (실적 + 예측)"
    ws["A1"].font = title_font
    ws.merge_cells("A1:H1")
    ws["A2"] = (
        f"기준: '{SHEET_HOLD}' 보유수량 · yfinance 배당이력 · 환율 {fx:,.0f}원/$  "
        f"| 연간 합계 {annual:,.0f}원 · 월평균 {annual/12:,.0f}원"
    )
    ws["A2"].font = Font(size=10, color="666666")
    ws.merge_cells("A2:H2")
    ws["A3"] = "※ 노란행=예측, 초록행=실적(또는 이미 지급된 배당). 커버드콜/미국ETF는 환율·지급월 추정 오차가 있을 수 있습니다."
    ws["A3"].font = Font(size=9, color="888888")

    # 월별 요약
    ws["A5"] = "① 월별 합계"
    ws["A5"].font = title_font
    for c, h in enumerate(["월", "실적", "예측", "합계", "구분"], 1):
        cell = ws.cell(6, c, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin
        cell.alignment = Alignment(horizontal="center")

    for m in range(1, 13):
        r = 6 + m
        a = monthly[m]["actual"]
        p = monthly[m]["predict"]
        t = a + p
        kind = "실적" if a > 0 and p == 0 else ("혼합" if a > 0 else "예측")
        fill = actual_fill if kind == "실적" else (predict_fill if kind == "예측" else PatternFill("solid", fgColor="DDEBF7"))
        ws.cell(r, 1, f"{m}월").border = thin
        ws.cell(r, 1).fill = fill
        for c, v in enumerate([a, p, t], 2):
            cell = ws.cell(r, c, v)
            cell.number_format = "#,##0"
            cell.border = thin
            cell.fill = fill
        ws.cell(r, 5, kind).border = thin
        ws.cell(r, 5).fill = fill

    ws.cell(20, 1, "합계").font = Font(bold=True)
    for c in range(2, 5):
        cell = ws.cell(20, c, f"=SUM({get_column_letter(c)}7:{get_column_letter(c)}18)")
        cell.number_format = "#,##0"
        cell.font = Font(bold=True)

    # bar chart
    chart = BarChart()
    chart.type = "col"
    chart.title = f"{year} 월별 배당·이자"
    chart.y_axis.title = "원"
    data = Reference(ws, min_col=4, min_row=6, max_row=18)
    cats = Reference(ws, min_col=1, min_row=7, max_row=18)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    chart.width = 15
    chart.height = 8
    ws.add_chart(chart, "G5")

    # Top contributors
    ws["A22"] = "② 연간 배당 기여 Top 15"
    ws["A22"].font = title_font
    for c, h in enumerate(["종목코드", "종목명", "주기", "연예상(원)", "배당률", "평가액"], 1):
        cell = ws.cell(23, c, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin

    top = sorted([d for d in detail if d["annual"] > 0], key=lambda x: -x["annual"])[:15]
    for i, d in enumerate(top):
        r = 24 + i
        vals = [d["ticker"], d["name"], d["freq"], d["annual"], d["yield"] / 100, d["val"]]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v)
            cell.border = thin
            if c in (4, 6):
                cell.number_format = "#,##0"
            if c == 5:
                cell.number_format = "0.00%"

    # 상세 내역
    detail_start = 42
    ws.cell(detail_start, 1, "③ 종목·월별 상세 (실적/예측)").font = title_font
    for c, h in enumerate(
        ["일자", "월", "종목코드", "종목명", "원화", "달러", "원화환산", "구분"], 1
    ):
        cell = ws.cell(detail_start + 1, c, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin
        cell.alignment = Alignment(horizontal="center")

    for i, rec in enumerate(records):
        r = detail_start + 2 + i
        tot = rec["krw"] + rec["usd"] * fx
        fill = predict_fill if rec["predict"] else actual_fill
        row = [
            rec["date"],
            rec["month"],
            rec["ticker"],
            rec["name"],
            rec["krw"] or None,
            rec["usd"] or None,
            tot,
            rec["kind"],
        ]
        for c, v in enumerate(row, 1):
            cell = ws.cell(r, c, v)
            cell.border = thin
            cell.fill = fill
            if c == 1:
                cell.number_format = "YYYY-MM-DD"
            if c in (5, 7):
                cell.number_format = "#,##0"
            if c == 6:
                cell.number_format = "0.0000"

    last_detail = detail_start + 1 + len(records)
    ws.auto_filter.ref = f"A{detail_start+1}:H{last_detail}"

    for col, w in {
        "A": 12,
        "B": 10,
        "C": 14,
        "D": 34,
        "E": 12,
        "F": 12,
        "G": 14,
        "H": 10,
    }.items():
        ws.column_dimensions[col].width = w

    wb.save(path)
    return annual, monthly


def run_dividend_tab(file_path: Path, year=YEAR):
    """현재가 반영된 파일 기준으로 마지막 탭 '월별배당'(연·월 배당 실적+예측) 생성."""
    print(f"\n=== 배당 탭: {file_path.name} ===")
    print("보유종목 로드...")
    holdings, fx = load_holdings(file_path)
    print(f"  {len(holdings)}행, 환율 {fx:,.0f}")
    print("배당 이력 수집·예측 중 (yfinance)...")
    records, detail, fx = build_records(holdings, fx, year=year)
    print(f"  레코드 {len(records)}건")
    annual, monthly = write_sheet(file_path, records, detail, fx, year=year)
    print(f"저장 → 탭 '{SHEET_OUT}' (마지막 시트)")
    print(f"연간 합계: {annual:,.0f}원 (월평균 {annual/12:,.0f}원)")
    for m in range(1, 13):
        a = monthly[m]["actual"]
        p = monthly[m]["predict"]
        tag = "실적" if a and not p else ("혼합" if a and p else "예측")
        print(f"  {m:2d}월: {a+p:>12,.0f}원  ({tag})")
    pred_n = sum(1 for r in records if r["predict"])
    print(f"실적 {len(records)-pred_n}건 / 예측 {pred_n}건")
    return annual


def main():
    import argparse

    ap = argparse.ArgumentParser(description="월별배당 탭 생성 (실적+예측)")
    ap.add_argument("file", nargs="?", default=str(DEFAULT_FILE))
    ap.add_argument("--year", type=int, default=YEAR)
    args = ap.parse_args()
    p = resolve_workbook(Path(args.file))
    run_dividend_tab(p, year=args.year)


if __name__ == "__main__":
    main()
