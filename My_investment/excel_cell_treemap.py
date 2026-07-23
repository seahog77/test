# -*- coding: utf-8 -*-
"""엑셀 시트 위 셀 색상 Finviz 스타일 트리맵 (병합 없음 · 외부 파일 없음)."""
from __future__ import annotations

import math

import squarify
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ACCT_COLORS = {
    "DC": "E2EFDA",
    "개인연금1": "DDEBF7",
    "개인연금2": "D6EAF8",
    "ISA": "FFF2CC",
    "일반계좌": "FCE4EC",
    "CMA": "FCE4D6",
    "IRP": "E4DFEC",
    "은행예금": "D9EAD3",
}
DEFAULT_FILL = "ECEFF1"
THIN = Side(style="thin", color="FFFFFF")


def _fill(hex6: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex6)


def _short(s: str, n: int = 12) -> str:
    s = str(s or "").strip()
    return s if len(s) <= n else s[: n - 1] + "…"


def draw_cell_treemap(
    ws,
    items: list[dict],
    *,
    top_row: int,
    left_col: int,
    width_cols: int,
    height_rows: int,
    title: str | None = None,
    title_col_span: int = 8,
    value_key: str = "val",
    label_key: str = "name",
    acct_key: str = "acct",
    max_items: int = 40,
):
    rows = [x for x in items if float(x.get(value_key) or 0) > 0]
    if not rows:
        return 0
    rows.sort(key=lambda x: -float(x[value_key]))
    if len(rows) > max_items:
        tail = rows[max_items - 1 :]
        rows = rows[: max_items - 1]
        rows.append(
            {
                label_key: "기타",
                value_key: sum(float(x[value_key]) for x in tail),
                acct_key: "",
            }
        )

    total = sum(float(x[value_key]) for x in rows)
    if total <= 0:
        return 0

    r0 = top_row
    c0 = left_col
    r1 = top_row + height_rows - 1
    c1 = left_col + width_cols - 1

    if title:
        ws.cell(r0 - 1, c0, title).font = Font(bold=True, size=11, color="1F4E79")

    for r in range(r0, r1 + 1):
        for c in range(c0, c1 + 1):
            cell = ws.cell(r, c)
            cell.value = None
            cell.fill = PatternFill()
            cell.border = Border()

    sizes = [float(x[value_key]) for x in rows]
    normed = squarify.normalize_sizes(sizes, width_cols, height_rows)
    rects = squarify.squarify(normed, 0, 0, width_cols, height_rows)

    for item, rect in zip(rows, rects):
        if isinstance(rect, dict):
            x, y, dx, dy = rect["x"], rect["y"], rect["dx"], rect["dy"]
        else:
            x, y, dx, dy = rect
        c_start = c0 + int(math.floor(x))
        c_end = c0 + int(math.ceil(x + dx)) - 1
        r_start = r0 + int(math.floor(y))
        r_end = r0 + int(math.ceil(y + dy)) - 1
        c_end = max(c_start, min(c_end, c1))
        r_end = max(r_start, min(r_end, r1))
        if c_end < c_start or r_end < r_start:
            continue

        acct = item.get(acct_key) or ""
        fill = _fill(ACCT_COLORS.get(acct, DEFAULT_FILL))
        pct = float(item[value_key]) / total * 100

        for r in range(r_start, r_end + 1):
            for c in range(c_start, c_end + 1):
                cell = ws.cell(r, c)
                cell.fill = fill
                cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

        label_cell = ws.cell(r_start, c_start)
        label_cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        name = item.get("label_formula") or _short(item.get(label_key, ""))
        if item.get("label_formula"):
            label_cell.value = name
        elif pct >= 5:
            label_cell.value = f"{name}\n{pct:.1f}%"
        elif pct >= 2:
            label_cell.value = f"{pct:.1f}%"
        elif (c_end - c_start + 1) >= 2:
            label_cell.value = name
        label_cell.font = Font(size=8, bold=True)

    col_w = max(2.0, min(4.5, 180 / width_cols))
    for c in range(c0, c1 + 1):
        ws.column_dimensions[get_column_letter(c)].width = col_w
    for r in range(r0, r1 + 1):
        ws.row_dimensions[r].height = 16

    return len(rows)
