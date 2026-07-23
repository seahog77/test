# -*- coding: utf-8 -*-
import pandas as pd
import sys
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding='utf-8')

F_OLD = r'c:\Users\seaho\My project\My_investment\investment.xlsx'
F_NEW = r'c:\Users\seaho\My project\My_investment\investment_0711.xlsx'


def read_holdings(path):
    df = pd.read_excel(path, sheet_name='3. 종목현황', header=None).iloc[8:]
    df.columns = ['acct', 'n', 'country', 'tic', 'name', 'qty', 'avg', 'avg_usd',
                  'px', 'px_usd', 'val', 'w', 'div', 'cum', 'ret', 'cat', 'a', 'b']
    for c in ['qty', 'avg', 'avg_usd', 'px', 'px_usd', 'val', 'div', 'cum', 'ret', 'n']:
        df[c] = pd.to_numeric(df[c], errors='coerce')
    df = df[df['val'].notna() & (df['val'] > 0)].copy()
    df['acct'] = df['acct'].fillna('').astype(str)
    df['tic'] = df['tic'].astype(str)
    df['name'] = df['name'].astype(str)
    df['uid'] = df['acct'] + '|' + df['tic'] + '|' + df['name'] + '|' + df['n'].astype(int).astype(str)
    return df


h1, h2 = read_holdings(F_OLD), read_holdings(F_NEW)
wb1 = load_workbook(F_OLD, data_only=True)
wb2 = load_workbook(F_NEW, data_only=True)

print('=' * 80)
print('  investment.xlsx  →  investment_0711.xlsx  변화 요약')
print('  (old: 2026-06-23 / new: 2026-07-11)')
print('=' * 80)

t1, t2 = h1['val'].sum(), h2['val'].sum()
print('\n[총괄]')
print('  보유 종목 행수: %d → %d (+%d)' % (len(h1), len(h2), len(h2) - len(h1)))
print('  총 평가액:      %s → %s' % (f'{t1:,.0f}', f'{t2:,.0f}'))
print('  순변동:         %+.0f원 (%.2f%%)' % (t2 - t1, (t2 / t1 - 1) * 100))

# merge on business key acct|tic|name
k = ['acct', 'tic', 'name']
m = h1[k + ['qty', 'avg', 'val', 'px']].merge(
    h2[k + ['qty', 'avg', 'val', 'px']],
    on=k, how='outer', suffixes=('_o', '_n'), indicator=True)

added = m[m['_merge'] == 'right_only']
removed = m[m['_merge'] == 'left_only']
both = m[m['_merge'] == 'both'].copy()
both['dq'] = both['qty_n'] - both['qty_o']
both['dval'] = both['val_n'] - both['val_o']
both['davg'] = both['avg_n'] - both['avg_o']
changed = both[(both['dq'].abs() > 0.01) | (both['davg'].abs() > 1) | (both['dval'].abs() > 1000)]

print('  신규 종목: %d  |  삭제: %d  |  수량·평단·평가 변경: %d' % (
    len(added), len(removed), len(changed)))

print('\n[계좌별 평가액]')
a1 = h1.groupby('acct')['val'].sum()
a2 = h2.groupby('acct')['val'].sum()
for acct in sorted(set(a1.index) | set(a2.index), key=str):
    v1, v2 = a1.get(acct, 0), a2.get(acct, 0)
    if abs(v2 - v1) > 10000 or acct in ('', 'DC', '개인연금1', '일반계좌', 'CMA', 'ISA', 'IRP'):
        print('  %-14s %12s → %12s  (%+.0f)' % (acct or '(미기재)', f'{v1:,.0f}', f'{v2:,.0f}', v2 - v1))

if len(added):
    print('\n[신규]')
    for _, r in added.sort_values('val_n', ascending=False).iterrows():
        print('  + [%s] %s (%s)  수량%.0f 평단%.0f  평가%s' % (
            r['acct'], r['name'][:30], r['tic'], r['qty_n'], r['avg_n'], f"{r['val_n']:,.0f}"))

if len(removed):
    print('\n[삭제·전량매도]')
    for _, r in removed.sort_values('val_o', ascending=False).iterrows():
        print('  - [%s] %s (%s)  평가%s' % (
            r['acct'], r['name'][:30], r['tic'], f"{r['val_o']:,.0f}"))

if len(changed):
    print('\n[변경 TOP 25]')
    for _, r in changed.reindex(changed['dval'].abs().sort_values(ascending=False).index).head(25).iterrows():
        parts = []
        if abs(r['dq']) > 0.01:
            parts.append('수량 %.0f→%.0f (%+.0f)' % (r['qty_o'], r['qty_n'], r['dq']))
        if abs(r['davg']) > 1:
            parts.append('평단 %.0f→%.0f' % (r['avg_o'], r['avg_n']))
        parts.append('평가 %+.0f' % r['dval'])
        print('  [%s] %s  %s' % (r['acct'], str(r['name'])[:28], ' | '.join(parts)))

# category
print('\n[카테고리별 평가액]')
c1 = h1.groupby('cat')['val'].sum()
c2 = h2.groupby('cat')['val'].sum()
for cat in sorted(set(c1.index) | set(c2.index), key=str):
    v1, v2 = c1.get(cat, 0), c2.get(cat, 0)
    if abs(v2 - v1) > 500000:
        print('  %-12s %12s → %12s  (%+.0f)' % (str(cat)[:12], f'{v1:,.0f}', f'{v2:,.0f}', v2 - v1))

# other sheets row counts
print('\n[기타 시트 행수]')
for sn in ['7. 배당내역', '6. 입금내역', '5. 계좌내역(누적)', '8.거래내역']:
    r1, r2 = wb1[sn].max_row, wb2[sn].max_row
    if r1 != r2:
        print('  %s: %d → %d' % (sn, r1, r2))
    else:
        print('  %s: 동일 (%d)' % (sn, r1))
