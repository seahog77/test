# -*- coding: utf-8 -*-
"""Combine MAIN + W portfolios and compute allocation / recent trends."""
from __future__ import annotations

import json
import re
import warnings
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
import pandas as pd
import yfinance as yf

warnings.filterwarnings("ignore")

BASE = Path(__file__).resolve().parent


def _num(v):
    """숫자 또는 '=123+456' / '=G49' 단순 수식을 평가. 실패 시 None."""
    if isinstance(v, (int, float)):
        return float(v)
    if not isinstance(v, str):
        return None
    s = v.strip()
    if not s:
        return None
    if s.startswith("="):
        expr = s[1:].strip().replace(",", "")
        # 셀 참조(=G49)는 호출측에서 해석
        if re.fullmatch(r"[A-Za-z]+\d+", expr):
            return None
        if re.fullmatch(r"[\d\.\+\-\*\s]+", expr):
            try:
                return float(eval(expr, {"__builtins__": {}}, {}))
            except Exception:
                return None
        return None
    try:
        return float(s.replace(",", ""))
    except Exception:
        return None


def _cell_ref_value(ws_f, ws_d, formula: str):
    """'=G49' 형태를 같은 시트의 값으로 해석."""
    if not isinstance(formula, str):
        return None
    m = re.fullmatch(r"=([A-Za-z]+)(\d+)", formula.strip())
    if not m:
        return None
    col = openpyxl.utils.column_index_from_string(m.group(1))
    row = int(m.group(2))
    v = ws_d.cell(row, col).value
    n = _num(v)
    if n is not None:
        return n
    v2 = ws_f.cell(row, col).value
    return _num(v2)


def load_holdings(path: Path, src: str):
    wb_f = openpyxl.load_workbook(path, data_only=False)
    wb_d = openpyxl.load_workbook(path, data_only=True)
    ws_f = wb_f["계좌현황"]
    ws_d = wb_d["계좌현황"]
    fx = ws_d.cell(5, 10).value
    if not isinstance(fx, (int, float)):
        fx = ws_f.cell(5, 10).value
    if not isinstance(fx, (int, float)):
        fx = 1480.0
    rows = []
    for r in range(9, 200):
        tic = ws_d.cell(r, 4).value or ws_f.cell(r, 4).value
        name = ws_d.cell(r, 5).value or ws_f.cell(r, 5).value
        if tic is None and name is None:
            if rows and r > rows[-1]["r"] + 3:
                break
            continue
        if name and "포트폴리오" in str(name):
            break
        if str(tic) == "포트폴리오":
            break
        acct = ws_d.cell(r, 1).value or ws_f.cell(r, 1).value or ""
        country = ws_d.cell(r, 3).value or ws_f.cell(r, 3).value or ""
        qty = ws_d.cell(r, 6).value
        if qty is None:
            qty = ws_f.cell(r, 6).value
        g_raw = ws_d.cell(r, 7).value
        if g_raw is None:
            g_raw = ws_f.cell(r, 7).value
        h_raw = ws_d.cell(r, 8).value
        if h_raw is None:
            h_raw = ws_f.cell(r, 8).value
        i_raw = ws_d.cell(r, 9).value
        if i_raw is None:
            i_raw = ws_f.cell(r, 9).value
        j_raw = ws_d.cell(r, 10).value
        if j_raw is None:
            j_raw = ws_f.cell(r, 10).value

        g = _num(g_raw)
        if g is None and isinstance(g_raw, str):
            g = _cell_ref_value(ws_f, ws_d, g_raw)
        h = _num(h_raw)
        if h is None and isinstance(h_raw, str):
            h = _cell_ref_value(ws_f, ws_d, h_raw)
        i = _num(i_raw)
        if i is None and isinstance(i_raw, str):
            i = _cell_ref_value(ws_f, ws_d, i_raw)
        j = _num(j_raw)
        if j is None and isinstance(j_raw, str):
            j = _cell_ref_value(ws_f, ws_d, j_raw)

        val = ws_d.cell(r, 11).value
        q = float(qty) if isinstance(qty, (int, float)) else 0.0
        tic_s = str(tic).strip() if tic else ""
        name_s = str(name).strip() if name else ""
        is_cash = tic_s == "현금" or any(
            k in name_s for k in ("예금", "예수금", "CMA", "MMF", "현금성", "새마을")
        )

        if not isinstance(val, (int, float)) or val <= 0:
            # openpyxl 저장 후 data_only 캐시가 비면 K/I 수식이 숫자로 안 나옴.
            # 현금은 평단(G)=평가액인 경우가 많아 G*수량으로 복구.
            if is_cash and g is not None:
                val = g * (q if q else 1.0)
            else:
                ip = float(i or 0)
                jp = float(j or 0)
                val = ip * q + jp * float(fx) * q
                # 미국 종목만 J가 있고 I가 비는 경우 등은 위에서 처리.
                # 그래도 0이면 G/H 원가 쪽으로 한 번 더 시도(현금성).
                if val <= 0 and g is not None:
                    val = g * (q if q else 1.0)

        cost = 0.0
        if g is not None:
            cost += g * q
        if h is not None:
            cost += h * float(fx) * q
        rows.append(
            {
                "r": r,
                "src": src,
                "acct": str(acct),
                "country": str(country or ""),
                "tic": tic_s,
                "name": name_s,
                "qty": q,
                "val": float(val or 0),
                "cost": cost,
            }
        )
    return rows, float(fx)


def classify(row):
    n, t = row["name"], row["tic"]
    if t in ("현금",) or any(
        k in n for k in ("예금", "예수금", "CMA", "MMF", "현금성", "새마을")
    ):
        return "현금·예금"
    if (
        "커버드콜" in n
        or ("타켓" in n and "콜" in n)
        or "위클리커버드" in n
        or "데일리커버드" in n
    ):
        if any(k in n for k in ("다우", "배당", "금융고배당")):
            return "커버드콜·배당소득"
        if any(k in n for k in ("테크", "AI", "나스닥")):
            return "커버드콜·성장"
        return "커버드콜·지수"
    if any(k in n for k in ("국채", "채권", "하이일드", "혼합50", "채권혼합")):
        return "채권·혼합"
    if "S&P500" in n or "S&P 500" in n:
        return "미국 S&P500"
    if "나스닥" in n:
        return "미국 나스닥"
    if "테크" in n or "AI" in n:
        return "미국 테크"
    if "비만" in n:
        return "테마(비만치료)"
    if t in ("CVX", "XOM", "WMT"):
        return "미국 개별주"
    if "배당" in n:
        return "배당 ETF"
    if re.fullmatch(r"\d+(\.0)?", t) or re.fullmatch(r"[0-9][0-9A-Za-z]{5}", t):
        return "한국 개별/기타"
    return "기타"


def to_yf(tic, name):
    t = str(tic).strip()
    if t in ("현금",) or any(k in name for k in ("예금", "예수금", "CMA", "MMF", "현금")):
        return None
    if re.fullmatch(r"[A-Z]{1,5}", t):
        return t
    if re.fullmatch(r"\d+(\.0)?", t):
        return f"{str(int(float(t))).zfill(6)}.KS"
    if re.fullmatch(r"[0-9][0-9A-Za-z]{5}", t):
        return f"{t}.KS"
    return None


def main():
    h1, _ = load_holdings(BASE / "investment_0723.xlsx", "MAIN")
    h2, _ = load_holdings(BASE / "investment_0723_W.xlsx", "W")
    allh = h1 + h2
    for r in allh:
        r["theme"] = classify(r)

    total = sum(x["val"] for x in allh)
    print("COMBINED", f"{total:,.0f}", "rows", len(allh))
    print("MAIN", f"{sum(x['val'] for x in h1):,.0f}", "W", f"{sum(x['val'] for x in h2):,.0f}")

    sa = defaultdict(float)
    for x in allh:
        sa[(x["src"], x["acct"])] += x["val"]
    th = defaultdict(float)
    for x in allh:
        th[x["theme"]] += x["val"]

    by_tic = defaultdict(lambda: {"val": 0, "name": "", "qty": 0, "cost": 0, "srcs": set()})
    for x in allh:
        k = x["tic"]
        by_tic[k]["val"] += x["val"]
        by_tic[k]["name"] = x["name"]
        by_tic[k]["qty"] += x["qty"]
        by_tic[k]["cost"] += x["cost"]
        by_tic[k]["srcs"].add(x["src"])

    top = sorted(by_tic.items(), key=lambda z: -z[1]["val"])[:25]
    for t, info in top[:12]:
        print(f"{t:8} {info['name'][:30]:30} {info['val']:12,.0f} {info['val']/total*100:5.1f}%")

    symbols = {}
    for t, info in by_tic.items():
        s = to_yf(t, info["name"])
        if s:
            symbols[t] = s

    need = [t for t, _ in top if t in symbols][:18]
    for x in h2:
        if x["tic"] in symbols and x["tic"] not in need:
            need.append(x["tic"])

    end = datetime.now()
    start = end - timedelta(days=120)
    print("Fetching", len(need), "tickers...")
    perf = {}
    for t in need:
        s = symbols[t]
        try:
            df = yf.download(
                s,
                start=start.strftime("%Y-%m-%d"),
                end=end.strftime("%Y-%m-%d"),
                progress=False,
                auto_adjust=True,
            )
            if df is None or len(df) < 5:
                print("skip", t, s)
                continue
            close = df["Close"]
            if hasattr(close, "columns"):
                close = close.iloc[:, 0]
            close = close.dropna()
            last = float(close.iloc[-1])

            def ret(days):
                tgt = close.index[-1] - pd.Timedelta(days=days)
                past = close[close.index <= tgt]
                if len(past) == 0:
                    past = close.iloc[:1]
                return (last / float(past.iloc[-1]) - 1) * 100

            weekly = close.resample("W-FRI").last().dropna().tail(12)
            series = [
                {"date": d.strftime("%Y-%m-%d"), "px": round(float(v), 2)}
                for d, v in weekly.items()
            ]
            base_px = series[0]["px"] if series else last
            series_n = [
                {"date": x["date"], "idx": round(x["px"] / base_px * 100, 2)}
                for x in series
            ]
            perf[t] = {
                "symbol": s,
                "name": by_tic[t]["name"],
                "val": by_tic[t]["val"],
                "last": last,
                "r1w": ret(7),
                "r1m": ret(30),
                "r3m": ret(90),
                "series": series_n,
            }
            print(
                t,
                f"1w={perf[t]['r1w']:.1f}",
                f"1m={perf[t]['r1m']:.1f}",
                f"3m={perf[t]['r3m']:.1f}",
            )
        except Exception as e:
            print("ERR", t, e)

    tic_shares = [info["val"] / total for info in by_tic.values()]
    hhi_tic = sum(s * s for s in tic_shares)

    # covered-call aggregate
    cc_val = sum(v for k, v in th.items() if k.startswith("커버드콜"))
    us_eq = sum(
        v
        for k, v in th.items()
        if k
        in (
            "미국 S&P500",
            "미국 나스닥",
            "미국 테크",
            "커버드콜·배당소득",
            "커버드콜·성장",
            "커버드콜·지수",
            "미국 개별주",
            "배당 ETF",
        )
    )

    out = {
        "asof": end.strftime("%Y-%m-%d"),
        "total": total,
        "main_total": sum(x["val"] for x in h1),
        "w_total": sum(x["val"] for x in h2),
        "n_main": len(h1),
        "n_w": len(h2),
        "n_tickers": len(by_tic),
        "hhi_tic": hhi_tic,
        "top1_share": top[0][1]["val"] / total if top else 0,
        "top5_share": sum(x[1]["val"] for x in top[:5]) / total,
        "cc_share": cc_val / total,
        "us_eq_share": us_eq / total,
        "cash_share": th.get("현금·예금", 0) / total,
        "bond_share": th.get("채권·혼합", 0) / total,
        "theme": [
            {"name": k, "val": v, "pct": v / total * 100}
            for k, v in sorted(th.items(), key=lambda z: -z[1])
        ],
        "acct": [
            {"src": a[0], "acct": a[1], "val": v, "pct": v / total * 100}
            for a, v in sorted(sa.items(), key=lambda z: -z[1])
        ],
        "top_tickers": [
            {
                "tic": t,
                "name": info["name"],
                "val": info["val"],
                "pct": info["val"] / total * 100,
                "cost": info["cost"],
                "pnl_pct": (
                    (info["val"] - info["cost"]) / info["cost"] * 100
                    if info["cost"] > 0
                    else None
                ),
                "srcs": sorted(info["srcs"]),
                "theme": next((x["theme"] for x in allh if x["tic"] == t), ""),
            }
            for t, info in top
        ],
        "perf": perf,
        "holdings": allh,
    }
    out_path = BASE / "_portfolio_analysis.json"
    out_path.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print("Wrote", out_path)
    print(
        "HHI",
        round(hhi_tic, 4),
        "top1",
        round(out["top1_share"] * 100, 1),
        "CC",
        round(out["cc_share"] * 100, 1),
    )


if __name__ == "__main__":
    main()
