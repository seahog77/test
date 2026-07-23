# -*- coding: utf-8 -*-
"""예측 배당금을 엑셀 7. 배당내역 시트에 기록"""
import pandas as pd
import yfinance as yf
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import date
from collections import defaultdict
import calendar
import sys
import warnings
import shutil
from pathlib import Path

warnings.filterwarnings('ignore')
sys.stdout.reconfigure(encoding='utf-8')

FILE = Path(r'c:\Users\seaho\My project\My_investment\investment.xlsx')
YEAR = 2026
FX = 1530
DATA_START_ROW = 5
MIN_AMOUNT = 1000  # 1,000원 미만 스킵

# 백업
backup = FILE.with_name('investment_backup_before_dividend.xlsx')
shutil.copy2(FILE, backup)
print(f'백업 생성: {backup.name}')

# ── 보유종목 로드 (티커별 수량 합산) ──
df = pd.read_excel(FILE, sheet_name='3. 종목현황', header=None)
data = df.iloc[8:].copy()
data.columns = ['계좌','번호','국가','티커','종목명','수량','평단가','_7','현재가','_9','평가액','투자비중','배당금','누적수익','총수익률','종목카테고리','_16','_17']
data['평가액'] = pd.to_numeric(data['평가액'], errors='coerce')
data['수량'] = pd.to_numeric(data['수량'], errors='coerce')
data = data[data['평가액'] > 0]

holdings = data.groupby(['티커', '종목명', '국가', '종목카테고리'], dropna=False).agg(
    수량=('수량', 'sum'), 평가액=('평가액', 'sum')
).reset_index()

cash_mask = holdings['종목카테고리'].isin(['현금', '예금', '예수금'])
cash_holdings = holdings[cash_mask]
stock_holdings = holdings[~cash_mask]


def to_yf_ticker(ticker, country):
    t = str(ticker).strip()
    if country == '미국':
        return t
    if t.isdigit():
        return f'{t.zfill(6)}.KS'
    return f'{t}.KS'


def clean_name(name):
    """계좌별 접미사 제거한 대표 종목명"""
    n = str(name)
    for suffix in ['(DC)', '(ISA)', '(개인연금)', '(삼성개인연금)', '(미래에셋IRP)', ', DC']:
        n = n.replace(suffix, '')
    return n.strip()


def typical_payday(year, month, sample_dates):
    """과거 배당일 패턴 기반 지급일 추정"""
    same_month = [d.day for d in sample_dates if d.year <= year and d.month == month]
    if same_month:
        day = round(sum(same_month) / len(same_month))
    else:
        day = 28 if month == 2 else 30 if month in (4, 6, 9, 11) else 31
        day = min(day, calendar.monthrange(year, month)[1])
    return date(year, month, day)


# ── 배당 레코드 생성 ──
records = []  # (date, ticker, name, krw, usd, is_predict)

for _, row in stock_holdings.iterrows():
    ticker = row['티커']
    qty = row['수량']
    country = row['국가']
    name = clean_name(row['종목명'])
    yf_t = to_yf_ticker(ticker, country)

    try:
        divs = yf.Ticker(yf_t).dividends
        divs.index = divs.index.tz_localize(None) if divs.index.tz else divs.index
    except Exception:
        divs = pd.Series(dtype=float)

    if len(divs) == 0:
        continue

    hist = divs[divs.index >= pd.Timestamp(f'{YEAR - 2}-01-01')]
    hist_df = hist.reset_index()
    hist_df.columns = ['date', 'amt']
    hist_df['month'] = hist_df['date'].dt.month
    month_pattern = hist_df.groupby('month')['amt'].mean()
    sample_dates = list(hist_df['date'].dt.date)

    # 최근 6개월 월평균 (per share)
    last6 = divs[divs.index >= pd.Timestamp(f'{YEAR}-01-01') - pd.DateOffset(months=6)]
    if len(last6) >= 2:
        l6 = last6.reset_index()
        l6.columns = ['date', 'amt']
        l6['mk'] = l6['date'].dt.to_period('M')
        recent_ps = l6.groupby('mk')['amt'].sum().mean()
        is_monthly = len(l6.groupby('mk')) >= 8
    else:
        recent_ps = last6.mean() if len(last6) else 0
        is_monthly = False

    for m in range(1, 13):
        month_start = pd.Timestamp(f'{YEAR}-{m:02d}-01')
        month_end = month_start + pd.DateOffset(months=1)
        actual = divs[(divs.index >= month_start) & (divs.index < month_end)]
        if len(actual) > 0:
            per_share = actual.sum()
            pay_date = actual.index[-1].date()
            is_predict = False
        elif m in month_pattern.index:
            per_share = month_pattern[m]
            pay_date = typical_payday(YEAR, m, sample_dates)
            is_predict = True
        elif is_monthly and recent_ps > 0:
            per_share = recent_ps
            pay_date = typical_payday(YEAR, m, sample_dates)
            is_predict = True
        else:
            continue

        if country == '미국':
            usd = round(per_share * qty, 4)
            krw = 0
            total_krw = usd * FX
        else:
            krw = round(per_share * qty)
            usd = 0
            total_krw = krw

        if total_krw < MIN_AMOUNT:
            continue

        records.append({
            'date': pay_date,
            'ticker': ticker,
            'name': name,
            'krw': krw,
            'usd': usd,
            'predict': is_predict,
            'month': m,
        })

# 현금/예금 월이자
for _, row in cash_holdings.iterrows():
    val = row['평가액']
    rate = 0.035 if 'RP' in str(row['종목명']) or row['종목카테고리'] == '현금' else 0.03
    monthly_int = round(val * rate / 12)
    if monthly_int < MIN_AMOUNT:
        continue
    for m in range(1, 13):
        day = calendar.monthrange(YEAR, m)[1]
        records.append({
            'date': date(YEAR, m, day),
            'ticker': row['티커'],
            'name': str(row['종목명']),
            'krw': monthly_int,
            'usd': 0,
            'predict': True,
            'month': m,
        })

# 티커+월 병합 (동일 종목 합산)
merged = {}
for r in records:
    key = (r['ticker'], r['month'])
    if key not in merged:
        merged[key] = r.copy()
    else:
        merged[key]['krw'] += r['krw']
        merged[key]['usd'] += r['usd']
        merged[key]['predict'] = merged[key]['predict'] and r['predict']

final_records = sorted(merged.values(), key=lambda x: (x['date'], str(x['ticker'])))
print(f'생성 레코드: {len(final_records)}건')

# ── 엑셀 쓰기 ──
wb = openpyxl.load_workbook(FILE)
ws = wb['7. 배당내역']

# 기존 수식 행 확보
formula_template = {
    3: ws.cell(DATA_START_ROW, 3).value,
    4: ws.cell(DATA_START_ROW, 4).value,
    5: ws.cell(DATA_START_ROW, 5).value,
    10: ws.cell(DATA_START_ROW, 10).value,
}
needed_rows = DATA_START_ROW + len(final_records)
current_formula_rows = 0
for r in range(DATA_START_ROW, ws.max_row + 1):
    if ws.cell(r, 3).value and str(ws.cell(r, 3).value).startswith('='):
        current_formula_rows = r
    else:
        break

# 수식 행 부족 시 복사
if needed_rows > current_formula_rows:
    for r in range(current_formula_rows + 1, needed_rows + 1):
        for col, tmpl in formula_template.items():
            val = str(tmpl)
            val = val.replace(str(DATA_START_ROW), str(r))
            ws.cell(r, col, value=val)

# 기존 입력 데이터 초기화 (수식열 제외)
for r in range(DATA_START_ROW, needed_rows + 1):
    for col in [2, 6, 7, 8, 9]:
        ws.cell(r, col, value=None)

# 데이터 입력
for i, rec in enumerate(final_records):
    r = DATA_START_ROW + i
    ws.cell(r, 2, value=rec['date'])          # B: 일자
    ws.cell(r, 6, value=rec['ticker'])        # F: 종목코드
    ws.cell(r, 7, value=rec['name'])          # G: 종목명
    ws.cell(r, 8, value=rec['krw'] if rec['krw'] else None)  # H: 원화
    ws.cell(r, 9, value=rec['usd'] if rec['usd'] else None)   # I: 외화

wb.save(FILE)

# 검증
monthly = defaultdict(float)
for rec in final_records:
    krw_total = rec['krw'] + rec['usd'] * FX
    monthly[rec['month']] += krw_total

print(f'\n저장 완료: {FILE.name}')
print(f'기록 기간: {YEAR}년 1~12월')
print(f'연간 합계(환율{FX}): {sum(monthly.values()):,.0f}원')
for m in range(1, 13):
    print(f'  {m:2d}월: {monthly[m]:>12,.0f}원')
predict_cnt = sum(1 for r in final_records if r['predict'])
actual_cnt = len(final_records) - predict_cnt
print(f'\n실제 기반: {actual_cnt}건 / 예측: {predict_cnt}건')
