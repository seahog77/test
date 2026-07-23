# -*- coding: utf-8 -*-
"""금액순 탭: 누적 비중 90%까지 상위 종목 표+차트 갱신."""
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

sys.stdout.reconfigure(encoding="utf-8")

SRC = Path(r"c:\Users\seaho\My project\My_investment\investment_0719.xlsx")
SHEET = "계좌현황"
AMOUNT_SHEET = "금액순"
ACCT_ORDER_FIXED = ["DC", "개인연금1", "개인연금2", "ISA", "일반계좌"]
HOLD_START = 9
CUM_TARGET = 0.90


def cached_number(v):
    if isinstance(v, (int, float)):
        return float(v)
    if not isinstance(v, str):
        return None
    m = re.search(r",\s*([-+]?\d+(?:\.\d+)?(?:[eE][-+]?\d+)?)\s*\)\s*$", v)
    return float(m.group(1)) if m else None


def cell_number(ws_data, ws_fml, r, c):
    v = ws_data.cell(r, c).value
    if isinstance(v, (int, float)):
        return float(v)
    return cached_number(ws_fml.cell(r, c).value) or 0.0


def main():
    wb_val = openpyxl.load_workbook(SRC, data_only=True)
    ws_val = wb_val[SHEET]
    wb = openpyxl.load_workbook(SRC, data_only=False)
    ws = wb[SHEET]

    hold_rows = []
    for r in range(HOLD_START, 200):
        d = ws_val.cell(r, 4).value
        e = ws_val.cell(r, 5).value
        if d is None and e is None:
            if hold_rows and r > hold_rows[-1] + 2:
                break
            continue
        if e and "포트폴리오" in str(e):
            break
        if d and str(d) == "포트폴리오":
            break
        hold_rows.append(r)

    infos = []
    acct_totals = defaultdict(float)
    for r in hold_rows:
        a = ws_val.cell(r, 1).value or "미지정"
        val = cell_number(ws_val, ws, r, 11)
        info = {
            "acct": a,
            "country": ws_val.cell(r, 3).value,
            "tic": ws_val.cell(r, 4).value,
            "name": ws_val.cell(r, 5).value,
            "qty": ws_val.cell(r, 6).value,
            "val": val,
            "pnl": cell_number(ws_val, ws, r, 13),
            "ret": ws_val.cell(r, 14).value,
        }
        infos.append(info)
        acct_totals[a] += val

    others = [a for a in acct_totals if a not in ACCT_ORDER_FIXED]
    acct_order = [a for a in ACCT_ORDER_FIXED if a in acct_totals] + sorted(
        others, key=lambda a: -acct_totals[a]
    )
    infos_sorted = sorted(
        infos, key=lambda x: (acct_order.index(x["acct"]), -x["val"])
    )
    total_val = sum(x["val"] for x in infos_sorted)
    print(f"total={total_val:,.0f} holdings={len(infos_sorted)}")

    by_ticker = defaultdict(
        lambda: {"val": 0.0, "name": "", "tic": "", "accts": set(), "_nv": -1.0}
    )
    for info in infos_sorted:
        tic = str(info["tic"]) if info["tic"] is not None else ""
        if isinstance(info["tic"], float) and info["tic"] == int(info["tic"]):
            tic = str(int(info["tic"]))
        name = str(info["name"]) if info["name"] else tic
        if tic in ("현금", "None", "") or any(
            name.startswith(p)
            for p in ("예수금", "현금", "CMA", "새마을", "삼성MMF", "미래에셋현금")
        ):
            key = f"CASH::{name}"
        else:
            key = f"TIC::{tic}"
        by_ticker[key]["val"] += info["val"]
        if info["val"] > by_ticker[key]["_nv"]:
            by_ticker[key]["name"] = name
            by_ticker[key]["tic"] = tic
            by_ticker[key]["_nv"] = info["val"]
        by_ticker[key]["accts"].add(info["acct"])

    ticker_sorted = sorted(by_ticker.items(), key=lambda x: -x[1]["val"])

    top = []
    cum = 0.0
    for key, v in ticker_sorted:
        top.append((key, v))
        cum += v["val"] / total_val if total_val else 0
        if cum >= CUM_TARGET:
            break
    other_val = total_val - sum(v["val"] for _, v in top)
    print(f"top count={len(top)} cum={cum * 100:.1f}% other={other_val:,.0f}")

    if AMOUNT_SHEET in wb.sheetnames:
        del wb[AMOUNT_SHEET]
    ws2 = wb.create_sheet(AMOUNT_SHEET, 1)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )
    acct_fills = {
        "DC": PatternFill("solid", fgColor="E2EFDA"),
        "개인연금1": PatternFill("solid", fgColor="DDEBF7"),
        "개인연금2": PatternFill("solid", fgColor="D6EAF8"),
        "ISA": PatternFill("solid", fgColor="FFF2CC"),
        "일반계좌": PatternFill("solid", fgColor="FCE4EC"),
        "CMA": PatternFill("solid", fgColor="FCE4D6"),
        "IRP": PatternFill("solid", fgColor="E4DFEC"),
        "은행예금": PatternFill("solid", fgColor="D9EAD3"),
    }
    title_font = Font(bold=True, size=14, color="1F4E79")

    ws2["A1"] = f"포트폴리오 상위 종목 비중 (누적 {CUM_TARGET:.0%}까지)"
    ws2["A1"].font = title_font
    ws2.merge_cells("A1:G1")
    ws2["A2"] = (
        f"총 평가액 {total_val:,.0f}원 · 상위 {len(top)}종목 누적 {cum * 100:.1f}% "
        f"+ 기타 {len(ticker_sorted) - len(top)}종목"
    )
    ws2["A2"].font = Font(size=10, color="666666")

    wh = ["순위", "종목코드", "종목명", "보유계좌", "평가액(원)", "비중", "누적비중"]
    for c, h in enumerate(wh, 1):
        cell = ws2.cell(4, c, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin

    pie_start = 5
    running = 0.0
    pie_end = pie_start
    for i, (key, v) in enumerate(top, 1):
        r = 4 + i
        w = v["val"] / total_val if total_val else 0
        running += w
        accts = ", ".join(
            sorted(
                v["accts"],
                key=lambda a: acct_order.index(a) if a in acct_order else 99,
            )
        )
        tic_disp = v.get("tic") or key.split("::", 1)[-1]
        row = [i, tic_disp, v["name"], accts, v["val"], w, running]
        for c, val in enumerate(row, 1):
            cell = ws2.cell(r, c, val)
            cell.border = thin
            if c == 5:
                cell.number_format = "#,##0"
                cell.font = Font(bold=True)
            if c in (6, 7):
                cell.number_format = "0.0%"
        pie_end = r

    other_row = pie_end + 1
    ws2.cell(other_row, 1, len(top) + 1).border = thin
    ws2.cell(other_row, 2, "기타").border = thin
    ws2.cell(other_row, 3, f"나머지 {len(ticker_sorted) - len(top)}종목").border = thin
    ws2.cell(other_row, 4, "-").border = thin
    cell = ws2.cell(other_row, 5, other_val)
    cell.number_format = "#,##0"
    cell.border = thin
    cell = ws2.cell(other_row, 6, other_val / total_val if total_val else 0)
    cell.number_format = "0.0%"
    cell.border = thin
    cell = ws2.cell(other_row, 7, 1.0)
    cell.number_format = "0.0%"
    cell.border = thin

    sum_row = other_row + 1
    ws2.cell(sum_row, 3, "합계").font = Font(bold=True)
    cell = ws2.cell(sum_row, 5, f"=SUM(E{pie_start}:E{other_row})")
    cell.number_format = "#,##0"
    cell.font = Font(bold=True)
    cell = ws2.cell(sum_row, 6, f"=SUM(F{pie_start}:F{other_row})")
    cell.number_format = "0.0%"
    cell.font = Font(bold=True)

    ws2.cell(4, 9, "차트라벨").font = Font(color="808080", size=8)
    ws2.cell(4, 10, "차트금액").font = Font(color="808080", size=8)
    for i, (key, v) in enumerate(top, 1):
        ws2.cell(4 + i, 9, str(v["name"])[:18])
        ws2.cell(4 + i, 10, v["val"])
    ws2.cell(other_row, 9, "기타")
    ws2.cell(other_row, 10, other_val)

    pie = PieChart()
    pie.title = f"누적 {CUM_TARGET:.0%} 상위 {len(top)}종목 + 기타"
    labels = Reference(ws2, min_col=9, min_row=pie_start, max_row=other_row)
    data = Reference(ws2, min_col=10, min_row=4, max_row=other_row)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showVal = False
    pie.dataLabels.showCatName = False
    pie.width = 18
    pie.height = 12
    ws2.add_chart(pie, "L4")

    acct_start = sum_row + 3
    ws2.cell(acct_start, 1, "계좌별 합계").font = title_font
    for c, h in enumerate(["계좌", "평가액(원)", "비중", "종목수"], 1):
        cell = ws2.cell(acct_start + 1, c, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin
    acct_counts = defaultdict(int)
    for info in infos_sorted:
        acct_counts[info["acct"]] += 1
    for j, a in enumerate(acct_order):
        rr = acct_start + 2 + j
        fill = acct_fills.get(a, PatternFill())
        ws2.cell(rr, 1, a).fill = fill
        ws2.cell(rr, 1).border = thin
        cell = ws2.cell(rr, 2, acct_totals[a])
        cell.number_format = "#,##0"
        cell.fill = fill
        cell.border = thin
        cell = ws2.cell(rr, 3, acct_totals[a] / total_val if total_val else 0)
        cell.number_format = "0.0%"
        cell.fill = fill
        cell.border = thin
        ws2.cell(rr, 4, acct_counts[a]).fill = fill
        ws2.cell(rr, 4).border = thin

    list_start = acct_start + 2 + len(acct_order) + 3
    ws2.cell(
        list_start,
        1,
        "전체 종목 (DC → 개인연금1 → 개인연금2 → ISA → 일반계좌 … / 계좌 내 금액↓)",
    ).font = title_font
    ws2.merge_cells(
        start_row=list_start, start_column=1, end_row=list_start, end_column=8
    )
    headers = ["순위", "계좌", "국가", "종목코드", "종목명", "수량", "평가액(원)", "투자비중"]
    hdr_row = list_start + 1
    for c, h in enumerate(headers, 1):
        cell = ws2.cell(hdr_row, c, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin

    for i, info in enumerate(infos_sorted, 1):
        r = hdr_row + i
        w = info["val"] / total_val if total_val else 0
        row = [
            i,
            info["acct"],
            info["country"],
            info["tic"],
            info["name"],
            info["qty"],
            info["val"],
            w,
        ]
        fill = acct_fills.get(info["acct"], PatternFill())
        for c, val in enumerate(row, 1):
            cell = ws2.cell(r, c, val)
            cell.border = thin
            cell.fill = fill
            if c in (6, 7):
                cell.number_format = "#,##0.##"
            if c == 8:
                cell.number_format = "0.00%"
            if c == 7:
                cell.font = Font(bold=True)

    last_list = hdr_row + len(infos_sorted)
    ws2.auto_filter.ref = f"A{hdr_row}:H{last_list}"
    ws2.freeze_panes = f"A{hdr_row + 1}"
    for col, w in {
        "A": 6,
        "B": 12,
        "C": 36,
        "D": 28,
        "E": 14,
        "F": 10,
        "G": 14,
        "H": 22,
        "I": 10,
    }.items():
        ws2.column_dimensions[col].width = w
    ws2.row_dimensions[4].height = 28

    wb.save(SRC)
    print(f"saved: {SRC}")
    print(f"\n=== 누적 {CUM_TARGET:.0%} 상위 {len(top)}종목 ===")
    running = 0.0
    for i, (key, v) in enumerate(top, 1):
        w = v["val"] / total_val * 100
        running += w
        name = v["name"][:32]
        print(f"{i:2d}. {name:32} {v['val']:>12,.0f} {w:5.1f}%  cum {running:5.1f}%")
    print(f"기타 {other_val:,.0f} {(other_val / total_val) * 100:.1f}%")


if __name__ == "__main__":
    main()
