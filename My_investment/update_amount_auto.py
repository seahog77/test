# -*- coding: utf-8 -*-
"""금액순 탭: 계좌현황 연동 수식 + 100% 전체 비중 표/차트."""
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app_paths import safe_reconfigure_stdio

safe_reconfigure_stdio()

LEFT = "계좌현황"
RIGHT = "금액순"
ACCT_ORDER = ["DC", "개인연금1", "개인연금2", "ISA", "일반계좌", "CMA", "IRP", "은행예금"]
HOLD_START = 9
DEFAULT_FILE = Path(r"c:\Users\seaho\My project\My_investment\investment_0719.xlsx")


def cached_number(v):
    if isinstance(v, (int, float)):
        return float(v)
    if not isinstance(v, str):
        return None
    m = re.search(r",\s*([-+]?\d+(?:\.\d+)?(?:[eE][-+]?\d+)?)\s*\)\s*$", v)
    return float(m.group(1)) if m else None


def cell_number(ws_d, ws_f, r, c):
    v = ws_d.cell(r, c).value
    if isinstance(v, (int, float)):
        return float(v)
    return cached_number(ws_f.cell(r, c).value) or 0.0


def detect_hold_rows(ws_d, ws_f):
    hold_rows = []
    for r in range(HOLD_START, 300):
        d = ws_d.cell(r, 4).value
        e = ws_d.cell(r, 5).value
        if d is None and e is None:
            d2 = ws_f.cell(r, 4).value
            e2 = ws_f.cell(r, 5).value
            if d2 is None and e2 is None:
                if hold_rows and r > hold_rows[-1] + 2:
                    break
                continue
        if (e and "포트폴리오" in str(e)) or (d and str(d) == "포트폴리오"):
            break
        tic = str(d or ws_f.cell(r, 4).value or "").strip()
        name = str(e or ws_f.cell(r, 5).value or "").strip()
        if not tic and not name:
            if hold_rows and r > hold_rows[-1] + 2:
                break
            continue
        hold_rows.append(r)
    return hold_rows


def update_amount_sheet(src: Path):
    wb_d = openpyxl.load_workbook(src, data_only=True)
    wb = openpyxl.load_workbook(src, data_only=False)
    ws_d = wb_d[LEFT]
    ws_l = wb[LEFT]

    hold_rows = detect_hold_rows(ws_d, ws_l)
    if not hold_rows:
        raise ValueError(f"'{LEFT}'에 종목 행이 없습니다.")

    first, last = hold_rows[0], hold_rows[-1]
    n = len(hold_rows)
    print(f"\n=== 금액순 탭: {src.name} ===")
    print(f"left range: {LEFT}!{first}:{last} ({n} rows)")

    # snapshot values (K 수식 미계산이면 I/J·수량으로 추정)
    fx = ws_d.cell(5, 10).value
    if not isinstance(fx, (int, float)):
        fx = 1480.0
    rows = []
    acct_totals = defaultdict(float)
    for r in hold_rows:
        a = ws_d.cell(r, 1).value or ws_l.cell(r, 1).value or ""
        val = cell_number(ws_d, ws_l, r, 11)
        if val <= 0:
            qty = cell_number(ws_d, ws_l, r, 6)
            i_px = cell_number(ws_d, ws_l, r, 9)
            j_px = cell_number(ws_d, ws_l, r, 10)
            if qty > 0 and (i_px > 0 or j_px > 0):
                val = i_px * qty + j_px * float(fx) * qty
        rows.append(
            {
                "r": r,
                "acct": a,
                "country": ws_d.cell(r, 3).value,
                "tic": ws_d.cell(r, 4).value or ws_l.cell(r, 4).value,
                "name": ws_d.cell(r, 5).value or ws_l.cell(r, 5).value,
                "qty": ws_d.cell(r, 6).value or ws_l.cell(r, 6).value,
                "val": val,
            }
        )
        acct_totals[a] += val
    total = sum(x["val"] for x in rows)
    print(f"total={total:,.0f}")

    if RIGHT in wb.sheetnames:
        del wb[RIGHT]
    ws = wb.create_sheet(RIGHT, 1)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )
    note_font = Font(size=10, color="666666")
    title_font = Font(bold=True, size=14, color="1F4E79")
    acct_fills = {
        "DC": PatternFill("solid", fgColor="E2EFDA"),
        "개인연금1": PatternFill("solid", fgColor="DDEBF7"),
        "개인연금2": PatternFill("solid", fgColor="D6EAF8"),
        "ISA": PatternFill("solid", fgColor="FFF2CC"),
        "일반계좌": PatternFill("solid", fgColor="FCE4EC"),
        "CMA": PatternFill("solid", fgColor="FCE4D6"),
        "IRP": PatternFill("solid", fgColor="E4DFEC"),
        "은행예금": PatternFill("solid", fgColor="D9EAD3"),
    }

    # ===== Header =====
    ws["A1"] = "금액순 · 포트폴리오 비중 100% (계좌현황 자동연동)"
    ws["A1"].font = title_font
    ws.merge_cells("A1:H1")
    ws["A2"] = (
        f"왼쪽 '{LEFT}' 시트의 평가액(K열)·계좌(A열)을 수식으로 참조합니다. "
        f"왼쪽을 수정·정렬하면 이 탭 수치가 함께 바뀝니다. "
        f"(Google 스프레드시트 / Excel 365 권장)"
    )
    ws["A2"].font = note_font
    ws.merge_cells("A2:H2")

    # Total via formula
    ws["A3"] = "총 평가액"
    ws["B3"] = f"=SUM('{LEFT}'!K{first}:K{last})"
    ws["B3"].number_format = "#,##0"
    ws["B3"].font = Font(bold=True, size=12, color="1F4E79")
    ws["C3"] = "종목수"
    ws["D3"] = f'=COUNTA(\'{LEFT}\'!E{first}:E{last})'
    ws["D3"].font = Font(bold=True)

    # ===== 1) 계좌별 합계 (SUMIF 자동) =====
    ws["A5"] = "① 계좌별 합계 (왼쪽 자동집계)"
    ws["A5"].font = title_font
    for c, h in enumerate(["계좌", "평가액", "비중", "종목수"], 1):
        cell = ws.cell(6, c, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin
        cell.alignment = Alignment(horizontal="center")

    # discover accounts present + fixed order
    present = [a for a in ACCT_ORDER if acct_totals.get(a, 0) or any(x["acct"] == a for x in rows)]
    for extra in acct_totals:
        if extra not in present:
            present.append(extra)

    for i, acct in enumerate(present):
        r = 7 + i
        fill = acct_fills.get(acct, PatternFill())
        ws.cell(r, 1, acct).fill = fill
        ws.cell(r, 1).border = thin
        # SUMIF from left sheet
        cell = ws.cell(r, 2, f"=SUMIF('{LEFT}'!A:A,A{r},'{LEFT}'!K:K)")
        cell.number_format = "#,##0"
        cell.fill = fill
        cell.border = thin
        cell = ws.cell(r, 3, f"=IF($B$3=0,0,B{r}/$B$3)")
        cell.number_format = "0.0%"
        cell.fill = fill
        cell.border = thin
        cell = ws.cell(
            r,
            4,
            f"=COUNTIF('{LEFT}'!A{first}:A{last},A{r})",
        )
        cell.fill = fill
        cell.border = thin

    acct_end = 6 + len(present)

    # ===== 2) 전체 종목 금액순 100% (수식 연동) =====
    list_title_row = acct_end + 3
    ws.cell(list_title_row, 1, "② 전체 종목 금액 내림차순 100% (왼쪽 K열 기준 자동정렬)").font = title_font
    ws.merge_cells(
        start_row=list_title_row,
        start_column=1,
        end_row=list_title_row,
        end_column=8,
    )

    hdr = list_title_row + 1
    headers = ["순위", "계좌", "국가", "종목코드", "종목명", "수량", "평가액", "비중", "누적비중"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(hdr, c, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Helper: unique key for duplicate values = rank with COUNTIF of larger values
    # For each output row i (1..n), find the i-th largest and its row on left sheet.
    # Use a robust approach with helper column on RIGHT sheet that mirrors left vals,
    # then INDEX/MATCH with tie-break via row number.
    #
    # Mirror block starting at column L (hidden-ish): row index + value from left
    # L = source row, M = value, N = tie-break key = value*100000 + (10000-row)
    # Mirror block: L=원본행, M=평가액, N=정렬키
    MCOL, VCOL, SCOL = 12, 13, 14
    mirror_start = hdr + 1
    for i, src_r in enumerate(hold_rows):
        rr = mirror_start + i
        ws.cell(rr, MCOL, src_r)
        ws.cell(rr, VCOL, f"='{LEFT}'!K{src_r}")
        ws.cell(
            rr,
            SCOL,
            f"=IF(M{rr}=\"\",-1E99,M{rr}*100000+(10000-L{rr}))",
        )

    mirror_end = mirror_start + n - 1
    ws.cell(hdr, MCOL, "원본행")
    ws.cell(hdr, VCOL, "평가액미러")
    ws.cell(hdr, SCOL, "정렬키")
    for c in (MCOL, VCOL, SCOL):
        ws.cell(hdr, c).font = Font(size=8, color="999999")

    # Data rows: for rank i, find LARGE of sort keys
    for i in range(1, n + 1):
        r = hdr + i
        # A: rank
        ws.cell(r, 1, i).border = thin
        # Find source row via MATCH on sort key
        # L_src = INDEX($L$..$L$, MATCH(LARGE($N$..$N$, i), $N$..$N$, 0))
        large = f"LARGE($N${mirror_start}:$N${mirror_end},{i})"
        idx = f"MATCH({large},$N${mirror_start}:$N${mirror_end},0)"
        row_ref = f"INDEX($L${mirror_start}:$L${mirror_end},{idx})"

        # B 계좌
        cell = ws.cell(r, 2, f"=IFERROR(INDEX('{LEFT}'!A:A,{row_ref}),\"\")")
        cell.border = thin
        # C 국가
        cell = ws.cell(r, 3, f"=IFERROR(INDEX('{LEFT}'!C:C,{row_ref}),\"\")")
        cell.border = thin
        # D 종목코드
        cell = ws.cell(r, 4, f"=IFERROR(INDEX('{LEFT}'!D:D,{row_ref}),\"\")")
        cell.border = thin
        # E 종목명
        cell = ws.cell(r, 5, f"=IFERROR(INDEX('{LEFT}'!E:E,{row_ref}),\"\")")
        cell.border = thin
        # F 수량
        cell = ws.cell(r, 6, f"=IFERROR(INDEX('{LEFT}'!F:F,{row_ref}),\"\")")
        cell.number_format = "#,##0.##"
        cell.border = thin
        # G 평가액
        cell = ws.cell(r, 7, f"=IFERROR(INDEX('{LEFT}'!K:K,{row_ref}),0)")
        cell.number_format = "#,##0"
        cell.font = Font(bold=True)
        cell.border = thin
        # H 비중
        cell = ws.cell(r, 8, f"=IF($B$3=0,0,G{r}/$B$3)")
        cell.number_format = "0.00%"
        cell.border = thin
        # I 누적비중
        if i == 1:
            cell = ws.cell(r, 9, f"=H{r}")
        else:
            cell = ws.cell(r, 9, f"=I{r-1}+H{r}")
        cell.number_format = "0.00%"
        cell.border = thin

    data_end = hdr + n

    # Seed cached-looking number formats; apply light alt rows using value snapshot order
    by_amt = sorted(rows, key=lambda x: -x["val"])
    for i, info in enumerate(by_amt, 1):
        r = hdr + i
        fill = acct_fills.get(info["acct"], PatternFill())
        for c in range(1, 10):
            if ws.cell(r, c).fill.fgColor is None or ws.cell(r, c).fill.fill_type is None:
                ws.cell(r, c).fill = fill

    # sum row
    sum_r = data_end + 1
    ws.cell(sum_r, 5, "합계").font = Font(bold=True)
    cell = ws.cell(sum_r, 7, f"=SUM(G{hdr+1}:G{data_end})")
    cell.number_format = "#,##0"
    cell.font = Font(bold=True)
    cell = ws.cell(sum_r, 8, f"=SUM(H{hdr+1}:H{data_end})")
    cell.number_format = "0.00%"
    cell.font = Font(bold=True)

    chart_block = sum_r + 3
    ws.cell(chart_block, 1, "③ 비중 차트 (전체 100%)").font = title_font
    ws.merge_cells(
        start_row=chart_block,
        start_column=1,
        end_row=chart_block,
        end_column=8,
    )

    def _no_labels(pie: PieChart) -> None:
        d = DataLabelList()
        d.showLegendKey = False
        d.showVal = False
        d.showCatName = False
        d.showSerName = False
        d.showPercent = False
        pie.series[0].dLbls = d

    pie_acct = PieChart()
    pie_acct.title = "계좌별"
    acct_labels = Reference(ws, min_col=1, min_row=7, max_row=acct_end)
    acct_data = Reference(ws, min_col=2, min_row=6, max_row=acct_end)
    pie_acct.add_data(acct_data, titles_from_data=True)
    pie_acct.set_categories(acct_labels)
    pie_acct.legend = None
    pie_acct.width = 14
    pie_acct.height = 10
    _no_labels(pie_acct)
    ws.add_chart(pie_acct, "K5")

    pie_hold = PieChart()
    pie_hold.title = "종목별"
    hold_labels = Reference(ws, min_col=5, min_row=hdr + 1, max_row=data_end)
    hold_data = Reference(ws, min_col=7, min_row=hdr, max_row=data_end)
    pie_hold.add_data(hold_data, titles_from_data=True)
    pie_hold.set_categories(hold_labels)
    pie_hold.legend = None
    pie_hold.width = 18
    pie_hold.height = 12
    _no_labels(pie_hold)
    ws.add_chart(pie_hold, f"K{chart_block + 1}")

    for col, w in {
        "A": 8,
        "B": 12,
        "C": 8,
        "D": 12,
        "E": 36,
        "F": 10,
        "G": 14,
        "H": 10,
        "I": 10,
        "L": 8,
        "M": 12,
        "N": 14,
    }.items():
        ws.column_dimensions[col].width = w

    ws.auto_filter.ref = f"A{hdr}:I{data_end}"
    ws.row_dimensions[hdr].height = 28
    ws.cell(hdr - 1, MCOL, "← 자동계산용(원본행/미러)").font = Font(
        size=8, color="AAAAAA"
    )

    wb.save(src)
    print(f"saved: {src}")
    print(f"amount-sorted rows: {n} (100%)")
    print("formulas link to", LEFT)
    # preview top 10 by value
    print("\n=== preview top 10 ===")
    for i, info in enumerate(by_amt[:10], 1):
        w = info["val"] / total * 100 if total else 0
        print(f"{i:2d}. [{info['acct']}] {str(info['name'])[:30]:30} {info['val']:>12,.0f} {w:5.1f}%")


def main():
    import argparse

    ap = argparse.ArgumentParser(description="금액순 탭 재생성 (종목 수 자동)")
    ap.add_argument("file", nargs="?", default=str(DEFAULT_FILE))
    args = ap.parse_args()
    p = Path(args.file)
    if not p.is_absolute():
        p = Path(__file__).resolve().parent / p
    update_amount_sheet(p)


if __name__ == "__main__":
    main()
