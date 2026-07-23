# -*- coding: utf-8 -*-
"""계좌 지정순 정렬 + 금액순 탭에 상위 종목 비중 표/차트."""
import re
import shutil
import sys
from collections import defaultdict
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
from openpyxl.formula.translate import Translator
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

SRC = Path(r"c:\Users\seaho\My project\My_investment\investment_0719.xlsx")
BAK = SRC.with_name("investment_0719_before_resort3.xlsx")
SHEET = "계좌현황"
AMOUNT_SHEET = "금액순"

# 사용자 지정 계좌 순서 (나머지는 뒤에 합계 큰 순)
ACCT_ORDER_FIXED = ["DC", "개인연금1", "개인연금2", "ISA", "일반계좌"]

HOLD_START = 9
HOLD_END = 120
MAX_COL = 18
TOP_N = 15  # 상위 종목 비중


def cached_number(formula_or_value):
    """data_only 값이 없을 때 IFERROR(..., cache) 형태에서 숫자 추출."""
    if isinstance(formula_or_value, (int, float)):
        return float(formula_or_value)
    if not isinstance(formula_or_value, str):
        return None
    m = re.search(r",\s*([-+]?\d+(?:\.\d+)?(?:[eE][-+]?\d+)?)\s*\)\s*$", formula_or_value)
    if m:
        try:
            return float(m.group(1))
        except ValueError:
            return None
    return None


def cell_number(ws_data, ws_fml, r, c):
    v = ws_data.cell(r, c).value
    if isinstance(v, (int, float)):
        return float(v)
    return cached_number(ws_fml.cell(r, c).value) or 0.0


def snapshot_row(ws, r):
    cells = []
    for c in range(1, MAX_COL + 1):
        cell = ws.cell(r, c)
        cells.append(
            {
                "value": cell.value,
                "number_format": cell.number_format,
                "font": copy(cell.font),
                "fill": copy(cell.fill),
                "border": copy(cell.border),
                "alignment": copy(cell.alignment),
                "protection": copy(cell.protection),
            }
        )
    return cells


def main():
    if not BAK.exists():
        shutil.copy2(SRC, BAK)
        print(f"backup: {BAK.name}")

    wb_val = openpyxl.load_workbook(SRC, data_only=True)
    ws_val = wb_val[SHEET]
    wb = openpyxl.load_workbook(SRC, data_only=False)
    ws = wb[SHEET]

    # find last holding row
    hold_rows = []
    for r in range(HOLD_START, HOLD_END + 1):
        d = ws_val.cell(r, 4).value
        e = ws_val.cell(r, 5).value
        if d is None and e is None:
            continue
        if e is not None and "포트폴리오" in str(e):
            break
        if d is not None and str(d) == "포트폴리오":
            break
        hold_rows.append(r)

    print(f"holding rows: {len(hold_rows)} ({hold_rows[0]}..{hold_rows[-1]})")

    acct_totals = defaultdict(float)
    row_vals = {}
    for r in hold_rows:
        a = ws_val.cell(r, 1).value
        # prefer data_only, fallback to formula cache
        k = cell_number(ws_val, ws, r, 11)
        key = a if a else "미지정"
        val = float(k)
        row_vals[r] = {
            "acct": key,
            "country": ws_val.cell(r, 3).value,
            "tic": ws_val.cell(r, 4).value,
            "name": ws_val.cell(r, 5).value,
            "qty": ws_val.cell(r, 6).value,
            "avg_krw": ws_val.cell(r, 7).value,
            "avg_usd": ws_val.cell(r, 8).value,
            "px_krw": cell_number(ws_val, ws, r, 9) or ws_val.cell(r, 9).value,
            "px_usd": cell_number(ws_val, ws, r, 10) or ws_val.cell(r, 10).value,
            "val": val,
            "weight": ws_val.cell(r, 12).value,
            "pnl": cell_number(ws_val, ws, r, 13),
            "ret": ws_val.cell(r, 14).value,
            "tag": ws_val.cell(r, 15).value,
        }
        acct_totals[key] += val

    # account order: fixed first, then others by total desc
    others = [a for a in acct_totals if a not in ACCT_ORDER_FIXED]
    others_sorted = sorted(others, key=lambda a: -acct_totals[a])
    acct_order = [a for a in ACCT_ORDER_FIXED if a in acct_totals] + others_sorted
    print("account order:", acct_order)
    for a in acct_order:
        print(f"  {a}: {acct_totals[a]:,.0f}")

    def sort_key(r):
        info = row_vals[r]
        return (acct_order.index(info["acct"]), -info["val"], r)

    sorted_rows = sorted(hold_rows, key=sort_key)
    snaps = {r: snapshot_row(ws, r) for r in hold_rows}

    clear_end = max(hold_rows[-1], HOLD_START + len(sorted_rows) - 1)
    for r in range(HOLD_START, clear_end + 1):
        for c in range(1, MAX_COL + 1):
            ws.cell(r, c).value = None

    new_order_info = []
    for i, old_r in enumerate(sorted_rows):
        new_r = HOLD_START + i
        snap = snaps[old_r]
        for c in range(1, MAX_COL + 1):
            src = snap[c - 1]
            dest = ws.cell(new_r, c)
            val = src["value"]
            if isinstance(val, str) and val.startswith("="):
                try:
                    val = Translator(
                        val, origin=f"{get_column_letter(c)}{old_r}"
                    ).translate_formula(f"{get_column_letter(c)}{new_r}")
                except Exception:
                    pass
            dest.value = val
            dest.number_format = src["number_format"]
            dest.font = src["font"]
            dest.fill = src["fill"]
            dest.border = src["border"]
            dest.alignment = src["alignment"]
            dest.protection = src["protection"]
        ws.cell(new_r, 2).value = i + 1
        if ws.cell(new_r, 1).value in (None, ""):
            ws.cell(new_r, 1).value = "미지정"
        info = dict(row_vals[old_r])
        info["new_r"] = new_r
        new_order_info.append(info)

    n_hold = len(sorted_rows)
    last_hold = HOLD_START + n_hold - 1
    print(f"rewrote {n_hold} rows -> {HOLD_START}..{last_hold}")

    ws.cell(8, 11).value = f"=SUM(K{HOLD_START}:K{last_hold})"
    for col in (13, 14):
        v = ws.cell(8, col).value
        if isinstance(v, str):
            m = re.match(r"(?i)=sum\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)", str(v))
            if m:
                ws.cell(8, col).value = (
                    f"=SUM({m.group(1)}{HOLD_START}:{m.group(3)}{last_hold})"
                )

    # ========== 금액순 sheet ==========
    if AMOUNT_SHEET in wb.sheetnames:
        del wb[AMOUNT_SHEET]
    ws2 = wb.create_sheet(AMOUNT_SHEET, 1)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )
    acct_fills = {
        "DC": PatternFill("solid", fgColor="E2EFDA"),
        "개인연금1": PatternFill("solid", fgColor="DDEBF7"),
        "개인연금2": PatternFill("solid", fgColor="D6EAF8"),
        "ISA": PatternFill("solid", fgColor="FFF2CC"),
        "일반계좌": PatternFill("solid", fgColor="FCE4EC"),
        "CMA": PatternFill("solid", fgColor="FCE4D6"),
        "IRP": PatternFill("solid", fgColor="E4DFEC"),
        "은행예금": PatternFill("solid", fgColor="D9EAD3"),
    }
    title_font = Font(bold=True, size=14, color="1F4E79")

    total_val = sum(x["val"] for x in new_order_info)

    # --- Section 1: 상위 종목 포트폴리오 비중 ---
    ws2["A1"] = "포트폴리오 상위 종목 비중"
    ws2["A1"].font = title_font
    ws2.merge_cells("A1:F1")
    ws2["A2"] = (
        f"총 평가액 {total_val:,.0f}원 · 상위 {TOP_N}개 + 기타 "
        f"(계좌순: {' → '.join(ACCT_ORDER_FIXED)} …)"
    )
    ws2["A2"].font = Font(size=10, color="666666")

    # Aggregate same ticker+name across accounts for portfolio view? 
    # User said 상위 종목 - use individual holdings (with account) as listed,
    # but for cleaner portfolio, aggregate by ticker across accounts.
    by_ticker = defaultdict(lambda: {"val": 0.0, "name": "", "tic": "", "accts": set(), "_name_val": -1.0})
    for info in new_order_info:
        tic = str(info["tic"]) if info["tic"] is not None else ""
        if isinstance(info["tic"], float) and info["tic"] == int(info["tic"]):
            tic = str(int(info["tic"]))
        name = str(info["name"]) if info["name"] else tic
        # 현금/예수금은 종목명별로 구분 (CMA·예금·MMF 섞이지 않게)
        if tic in ("현금", "None", "") or any(
            name.startswith(p) for p in ("예수금", "현금", "CMA", "새마을", "삼성MMF", "미래에셋현금")
        ):
            key = f"CASH::{name}"
        else:
            key = f"TIC::{tic}"
        by_ticker[key]["val"] += info["val"]
        if info["val"] > by_ticker[key]["_name_val"]:
            by_ticker[key]["name"] = name
            by_ticker[key]["tic"] = tic
            by_ticker[key]["_name_val"] = info["val"]
        by_ticker[key]["accts"].add(info["acct"])

    ticker_sorted = sorted(by_ticker.items(), key=lambda x: -x[1]["val"])
    top = ticker_sorted[:TOP_N]
    other_val = sum(v["val"] for _, v in ticker_sorted[TOP_N:])

    # weight table headers
    wh = ["순위", "종목코드", "종목명", "보유계좌", "평가액(원)", "비중"]
    for c, h in enumerate(wh, 1):
        cell = ws2.cell(4, c, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin

    pie_start = 5
    for i, (key, v) in enumerate(top, 1):
        r = 4 + i
        w = v["val"] / total_val if total_val else 0
        accts = ", ".join(
            sorted(v["accts"], key=lambda a: acct_order.index(a) if a in acct_order else 99)
        )
        name = v["name"]
        tic_disp = v.get("tic") or key.split("::", 1)[-1]
        row = [i, tic_disp, name, accts, v["val"], w]
        for c, val in enumerate(row, 1):
            cell = ws2.cell(r, c, val)
            cell.border = thin
            if c == 5:
                cell.number_format = "#,##0"
                cell.font = Font(bold=True)
            if c == 6:
                cell.number_format = "0.0%"
        pie_end = r

    # 기타 row
    other_row = pie_end + 1
    ws2.cell(other_row, 1, TOP_N + 1).border = thin
    ws2.cell(other_row, 2, "기타").border = thin
    ws2.cell(other_row, 3, f"나머지 {len(ticker_sorted) - TOP_N}종목").border = thin
    ws2.cell(other_row, 4, "-").border = thin
    cell = ws2.cell(other_row, 5, other_val)
    cell.number_format = "#,##0"
    cell.border = thin
    cell = ws2.cell(other_row, 6, other_val / total_val if total_val else 0)
    cell.number_format = "0.0%"
    cell.border = thin

    sum_row = other_row + 1
    ws2.cell(sum_row, 3, "합계").font = Font(bold=True)
    cell = ws2.cell(sum_row, 5, f"=SUM(E{pie_start}:E{other_row})")
    cell.number_format = "#,##0"
    cell.font = Font(bold=True)
    cell = ws2.cell(sum_row, 6, f"=SUM(F{pie_start}:F{other_row})")
    cell.number_format = "0.0%"
    cell.font = Font(bold=True)

    # Pie chart data for chart (labels in B, values in E) - use short labels in col H/I for chart
    ws2.cell(4, 8, "차트라벨")
    ws2.cell(4, 9, "차트금액")
    ws2.cell(4, 8).font = Font(color="808080", size=8)
    ws2.cell(4, 9).font = Font(color="808080", size=8)
    for i, (key, v) in enumerate(top, 1):
        label = str(v["name"])[:18]
        ws2.cell(4 + i, 8, label)
        ws2.cell(4 + i, 9, v["val"])
    ws2.cell(other_row, 8, "기타")
    ws2.cell(other_row, 9, other_val)

    pie = PieChart()
    pie.title = f"상위 {TOP_N}종목 + 기타 비중"
    labels = Reference(ws2, min_col=8, min_row=pie_start, max_row=other_row)
    data = Reference(ws2, min_col=9, min_row=4, max_row=other_row)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showVal = False
    pie.dataLabels.showCatName = False
    pie.width = 18
    pie.height = 12
    ws2.add_chart(pie, "K4")

    # --- Section 2: 계좌별 합계 ---
    acct_start = sum_row + 3
    ws2.cell(acct_start, 1, "계좌별 합계").font = title_font
    for c, h in enumerate(["계좌", "평가액(원)", "비중", "종목수"], 1):
        cell = ws2.cell(acct_start + 1, c, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin

    acct_counts = defaultdict(int)
    for info in new_order_info:
        acct_counts[info["acct"]] += 1

    for j, a in enumerate(acct_order):
        rr = acct_start + 2 + j
        fill = acct_fills.get(a, PatternFill())
        ws2.cell(rr, 1, a).fill = fill
        ws2.cell(rr, 1).border = thin
        cell = ws2.cell(rr, 2, acct_totals[a])
        cell.number_format = "#,##0"
        cell.fill = fill
        cell.border = thin
        cell = ws2.cell(rr, 3, acct_totals[a] / total_val if total_val else 0)
        cell.number_format = "0.0%"
        cell.fill = fill
        cell.border = thin
        ws2.cell(rr, 4, acct_counts[a]).fill = fill
        ws2.cell(rr, 4).border = thin

    # --- Section 3: 계좌순 + 계좌내 금액순 전체 목록 (뒤쪽 탭 본문) ---
    list_start = acct_start + 2 + len(acct_order) + 3
    ws2.cell(list_start, 1, "전체 종목 (계좌순 · 계좌 내 금액 내림차순)").font = title_font
    ws2.merge_cells(start_row=list_start, start_column=1, end_row=list_start, end_column=10)

    headers = [
        "순위",
        "계좌",
        "국가",
        "종목코드",
        "종목명",
        "수량",
        "평가액(원)",
        "투자비중",
        "누적수익(원)",
        "총수익률",
    ]
    hdr_row = list_start + 1
    for c, h in enumerate(headers, 1):
        cell = ws2.cell(hdr_row, c, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin

    # same order as 계좌현황 (account order, amount desc within)
    for i, info in enumerate(new_order_info, 1):
        r = hdr_row + i
        w = info["val"] / total_val if total_val else 0
        row = [
            i,
            info["acct"],
            info["country"],
            info["tic"],
            info["name"],
            info["qty"],
            info["val"],
            w,
            info["pnl"],
            info["ret"],
        ]
        fill = acct_fills.get(info["acct"], PatternFill())
        for c, val in enumerate(row, 1):
            cell = ws2.cell(r, c, val)
            cell.border = thin
            cell.fill = fill
            if c in (6, 7, 9):
                cell.number_format = "#,##0.##"
            if c in (8, 10) and isinstance(val, (int, float)):
                cell.number_format = "0.00%"
            if c == 7:
                cell.font = Font(bold=True)

    last_list = hdr_row + len(new_order_info)
    ws2.auto_filter.ref = f"A{hdr_row}:J{last_list}"
    ws2.freeze_panes = f"A{hdr_row + 1}"

    widths = {
        "A": 6,
        "B": 12,
        "C": 36,
        "D": 28,
        "E": 14,
        "F": 10,
        "G": 14,
        "H": 22,
        "I": 12,
        "J": 10,
    }
    for col, w in widths.items():
        ws2.column_dimensions[col].width = w
    ws2.column_dimensions["K"].width = 3
    ws2.row_dimensions[4].height = 28

    wb.save(SRC)
    print(f"saved: {SRC}")
    print(f"sheet '{AMOUNT_SHEET}' updated with top-{TOP_N} weights + pie chart")

    print("\n=== 계좌현황 순서 확인 ===")
    cur = None
    for info in new_order_info:
        if info["acct"] != cur:
            cur = info["acct"]
            print(f"\n[{cur}] {acct_totals[cur]:,.0f}")
        print(f"  {info['new_r']:3d} {str(info['name'])[:30]:30} {info['val']:,.0f}")

    print(f"\n=== 상위 {TOP_N} 비중 ===")
    if total_val <= 0:
        print("  (평가액 없음 — 캐시값 확인 필요)")
    else:
        cum = 0
        for i, (key, v) in enumerate(top, 1):
            w = v["val"] / total_val * 100
            cum += w
            print(f"  {i:2d}. {v['name'][:28]:28} {v['val']:>12,.0f}  {w:5.1f}%")
        print(f"  기타 {other_val:>12,.0f}  {other_val/total_val*100:5.1f}%")
        print(f"  상위{TOP_N} 누적 {cum:.1f}%")


if __name__ == "__main__":
    main()
